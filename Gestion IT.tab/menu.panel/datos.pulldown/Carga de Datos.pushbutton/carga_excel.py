# -*- coding: utf-8 -*-
"""
Script CPython — Lee hoja CODIGO del .xlsm y genera TXT temporal con filas.
Uso: python carga_excel.py <ruta_xlsm> <ruta_repo_tmp>
"""
import os
import sys
from openpyxl import load_workbook


def read_codigos_sheet(xlsm_path):
    try:
        wb = load_workbook(xlsm_path, data_only=True)
    except Exception as e:
        print(u"ERROR: No se pudo abrir el archivo Excel:\n{}\n".format(e))
        return []

    if "CODIGO" not in wb.sheetnames:
        print(u"ERROR: No se encontró la hoja 'CODIGO' en el archivo Excel.")
        return []

    sheet = wb["CODIGO"]
    filas = []
    try:
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            filas.append(row)
    except Exception as e:
        print(u"ERROR leyendo la hoja CODIGO:\n{}".format(e))
        return []
    return filas


def main():
    if len(sys.argv) < 3:
        print(u"Uso: carga_excel.py <ruta_xlsm> <ruta_repo_tmp>")
        sys.exit(1)

    xlsm_path    = sys.argv[1]
    repo_tmp_path = sys.argv[2]

    if not os.path.exists(xlsm_path):
        print(u"ERROR: No existe el archivo Excel:\n{}".format(xlsm_path))
        sys.exit(1)

    filas = read_codigos_sheet(xlsm_path)
    if not filas:
        print(u"AVISO: No se leyeron filas en 'CODIGO' (desde la fila 2).")
        try:
            with open(repo_tmp_path, "w", encoding="utf-8") as f:
                f.write("")
        except Exception as e:
            print(u"ERROR creando archivo temporal:\n{}".format(e))
            sys.exit(1)
        sys.exit(0)

    try:
        with open(repo_tmp_path, "w", encoding="utf-8") as f:
            for row in filas:
                valores = ["" if v is None else str(v) for v in row]
                f.write(";".join(valores) + "\n")
    except Exception as e:
        print(u"ERROR escribiendo repositorio temporal:\n{}".format(e))
        sys.exit(1)

    print(u"Datos escritos en:\n{}".format(repo_tmp_path))
    sys.exit(0)


if __name__ == "__main__":
    main()
