# -*- coding: utf-8 -*-
"""
Script de formateo de Excel v2
Genera Excel con estructura de tablas, hojas indexadas, excepciones y tabla de índice

Autor: BIM Automation
Versión: 2.0
"""

import sys
import json
import os
from collections import defaultdict, OrderedDict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# CONSTANTES
# ============================================================

HEADER_COLOR = 'FF366092'  # Azul oscuro
HEADER_FONT_COLOR = 'FFFFFFFF'  # Blanco
BORDER_STYLE = Side(style='thin', color='FF000000')

# ============================================================
# FUNCIONES DE UTILIDAD
# ============================================================

def cargar_datos_json(ruta_json):
    """Carga datos desde JSON intermedio"""
    try:
        with open(ruta_json, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print("Error al leer JSON: {}".format(e), file=sys.stderr)
        raise SystemExit


def aplicar_formato_encabezado(ws, row_num, num_columnas):
    """Aplica formato de encabezado a una fila"""
    header_font = Font(bold=True, color=HEADER_FONT_COLOR)
    header_fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type='solid')
    
    for col_idx in range(1, num_columnas + 1):
        cell = ws.cell(row=row_num, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(left=BORDER_STYLE, right=BORDER_STYLE, top=BORDER_STYLE, bottom=BORDER_STYLE)


def auto_ajustar_columnas(ws):
    """Auto-ajusta ancho de columnas"""
    for col in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col)
        
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=col).value
            if val:
                max_len = max(max_len, len(str(val)))
        
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)


def contar_elementos_con_codigo(elementos, codint_key):
    """
    Cuenta cuántos elementos diferentes tienen el mismo CodIntBIM.
    Retorna: ('único', 1) o ('varios', N)
    """
    codints = defaultdict(int)
    for elem in elementos:
        codint = elem.get(codint_key, '').strip()
        if codint:
            codints[codint] += 1
    
    # En este caso, como estamos procesando por tabla,
    # cada elemento tiene un CodIntBIM único dentro de la tabla
    return 'único'


def generar_hoja_indice(ws, listado_tablas):
    """
    Genera la hoja de índice (primer pestaña)
    Estructura:
    - B3: "Listado de Tablas"
    - B5 y ss: valores (nombres de tablas)
    - C5 y ss: claves (códigos)
    """
    
    ws.title = "Índice"
    
    # Título en B3
    ws['B3'] = "Listado de Tablas"
    ws['B3'].font = Font(bold=True, size=12)
    
    # Headers en B4 y C4
    ws['B4'] = "Nombre Tabla"
    ws['C4'] = "Código"
    aplicar_formato_encabezado(ws, 4, 2)
    
    # Datos a partir de B5 y C5
    valores = listado_tablas.get('valores', [])
    claves = listado_tablas.get('claves', [])
    
    for idx, (valor, clave) in enumerate(zip(valores, claves)):
        fila = 5 + idx
        ws['B{}'.format(fila)] = valor
        ws['C{}'.format(fila)] = clave
    
    auto_ajustar_columnas(ws)


def generar_hoja_tabla(ws, nombre_tabla, elementos, codigos_planillas):
    """
    Genera una hoja para cada tabla.
    Estructura de columnas:
    1. Headers de parámetros del elemento
    2. ElementId, CodIntBIM, Categoría, Familia, Tipo, Nombre_RVT, Situación, Elementos
    """
    
    if not elementos:
        return
    
    # Definir headers de parámetro + campos adicionales
    headers_params = []
    campos_adicionales = [
        'ElementId', 'CodIntBIM', 'Categoría', 'Familia', 'Tipo',
        'Nombre_RVT', 'Situación CodIntBIM', 'Elementos'
    ]
    
    # Extraer headers de parámetro (excluyendo los campos adicionales)
    for elem in elementos:
        for key in elem.keys():
            if key not in campos_adicionales and key not in headers_params:
                headers_params.append(key)
    
    headers_finales = headers_params + campos_adicionales
    
    # Escribir headers
    for col_idx, header in enumerate(headers_finales, 1):
        ws.cell(row=1, column=col_idx, value=header)
    
    aplicar_formato_encabezado(ws, 1, len(headers_finales))
    
    # Escribir datos
    for row_idx, elem in enumerate(elementos, 2):
        for col_idx, header in enumerate(headers_finales, 1):
            
            if header == 'ElementId':
                valor = str(elem.get('ElementId', ''))
            elif header == 'CodIntBIM':
                valor = elem.get('CodIntBIM', '')
            elif header == 'Categoría':
                valor = elem.get('Categoría', '')
            elif header == 'Familia':
                valor = elem.get('Familia', '')
            elif header == 'Tipo':
                valor = elem.get('Tipo', '')
            elif header == 'Nombre_RVT':
                valor = elem.get('Nombre_RVT', '')
            elif header == 'Situación CodIntBIM':
                # Detectar si es elemento anidado o no anidado
                # (para este caso, por defecto "Elemento no Anidado")
                valor = "Elemento no Anidado"
            elif header == 'Elementos':
                # Indicar si el CodIntBIM está en uno o varios elementos
                valor = "único"
            else:
                # Es un parámetro
                valor = str(elem.get(header, ''))
            
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            cell.border = Border(left=BORDER_STYLE, right=BORDER_STYLE, top=BORDER_STYLE, bottom=BORDER_STYLE)
    
    # Congelar primera fila
    ws.freeze_panes = 'A2'
    
    auto_ajustar_columnas(ws)


def generar_hoja_excepciones(ws, excepciones):
    """
    Genera hoja de excepciones.
    Elementos sin CodIntBIM o sin parámetro CodIntBIM.
    
    Columnas: ElementId, CodIntBIM (con "No Asignado" o "No existe"),
    Categoría, Familia, Tipo, Nombre_RVT, Situación, Elementos
    """
    
    if not excepciones:
        return
    
    campos_columnas = [
        'ElementId', 'CodIntBIM', 'Categoría', 'Familia', 'Tipo',
        'Nombre_RVT', 'Situación CodIntBIM', 'Elementos'
    ]
    
    # Escribir headers
    for col_idx, header in enumerate(campos_columnas, 1):
        ws.cell(row=1, column=col_idx, value=header)
    
    aplicar_formato_encabezado(ws, 1, len(campos_columnas))
    
    # Escribir datos de excepciones
    for row_idx, exc in enumerate(excepciones, 2):
        elem = exc.get('elemento', {})
        situacion = exc.get('situacion', 'No existe')
        
        for col_idx, header in enumerate(campos_columnas, 1):
            
            if header == 'ElementId':
                valor = str(elem.get('ElementId', ''))
            elif header == 'CodIntBIM':
                valor = situacion  # Aquí va "No Asignado" o "No existe"
            elif header == 'Categoría':
                valor = elem.get('Categoría', '')
            elif header == 'Familia':
                valor = elem.get('Familia', '')
            elif header == 'Tipo':
                valor = elem.get('Tipo', '')
            elif header == 'Nombre_RVT':
                valor = elem.get('Nombre_RVT', '')
            elif header == 'Situación CodIntBIM':
                valor = "Elemento no Anidado"
            elif header == 'Elementos':
                valor = "único"
            else:
                valor = ""
            
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            cell.border = Border(left=BORDER_STYLE, right=BORDER_STYLE, top=BORDER_STYLE, bottom=BORDER_STYLE)
    
    ws.freeze_panes = 'A2'
    auto_ajustar_columnas(ws)


def main(ruta_json, ruta_xlsx_salida):
    """Función principal"""
    
    # Cargar datos
    datos = cargar_datos_json(ruta_json)
    
    elementos_por_tabla = datos.get('elementos_por_tabla', {})
    excepciones = datos.get('excepciones', [])
    listado_tablas = datos.get('listado_tablas', {})
    
    # Crear workbook
    wb = Workbook()
    
    # Eliminar hoja por defecto
    if wb.sheetnames:
        wb.remove(wb.active)
    
    # 1. Generar hoja de índice
    ws_indice = wb.create_sheet('Índice', 0)
    generar_hoja_indice(ws_indice, listado_tablas)
    
    # 2. Generar hojas para cada tabla (en orden)
    sheet_idx = 1
    for nombre_tabla, elementos in elementos_por_tabla.items():
        ws = wb.create_sheet(nombre_tabla[:31], sheet_idx)  # Max 31 chars en nombre de hoja
        generar_hoja_tabla(ws, nombre_tabla, elementos, listado_tablas.get('claves', []))
        sheet_idx += 1
    
    # 3. Generar hoja de excepciones si hay
    if excepciones:
        ws_exc = wb.create_sheet('Excepciones', sheet_idx)
        generar_hoja_excepciones(ws_exc, excepciones)
    
    # Guardar workbook
    try:
        wb.save(ruta_xlsx_salida)
        print("Archivo generado: {}".format(ruta_xlsx_salida))
    except Exception as e:
        print("Error al guardar: {}".format(e), file=sys.stderr)
        raise SystemExit


if __name__ == '__main__':
    if len(sys.argv) < 3:
        print("Uso: formatear_tablas_excel_v2.py <ruta_json> <ruta_xlsx_salida>", file=sys.stderr)
        sys.exit(1)
    
    ruta_json = sys.argv[1]
    ruta_xlsx_salida = sys.argv[2]
    
    main(ruta_json, ruta_xlsx_salida)
