# -*- coding: utf-8 -*-
import sys
import os
import json
from datetime import datetime
from copy import copy

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Carpeta base donde estará script.json
BASE_DATA_DIR = os.path.join(
    os.path.expanduser("~"),
    r"AppData\Roaming\MyPyRevitExtention\PyRevitIT.extension\data"
)

def get_script_json_path():
    """Devuelve la ruta completa a script.json en la carpeta de datos."""
    return os.path.join(BASE_DATA_DIR, "script.json")


def cargar_config_json(json_path=None):
    """Carga reemplazos_encabezados y reemplazos_de_nombres desde script.json."""
    if json_path is None:
        json_path = get_script_json_path()
    with open(json_path, "r", encoding="utf-8") as f:
        config = json.load(f)
    return config.get("reemplazos_encabezados", {}), config.get("reemplazos_de_nombres", {})


def copiar_fila_formato(ws, fila_origen, fila_destino, max_col):
    for col in range(1, max_col + 1):
        origen = ws.cell(row=fila_origen, column=col)
        destino = ws.cell(row=fila_destino, column=col)
        destino.font = copy(origen.font)
        destino.border = copy(origen.border)
        destino.fill = copy(origen.fill)
        destino.alignment = copy(origen.alignment)
        destino.number_format = copy(origen.number_format)
        destino.protection = copy(origen.protection)


def procesar_encabezados_divididos(ws, fila_encabezado, reemplazos_encabezados):
    max_col = ws.max_column

    # Preparar títulos segmentados reemplazando según diccionario
    max_segmentos = 1
    titulos_segmentados = []

    for col in range(1, max_col + 1):
        celda = ws.cell(row=fila_encabezado, column=col)
        val = celda.value if celda.value else ""
        if val in reemplazos_encabezados:
            val = reemplazos_encabezados[val]
        segmentos = val.split('/')
        titulos_segmentados.append(segmentos)
        if len(segmentos) > max_segmentos:
            max_segmentos = len(segmentos)

    # Insertar filas hacia abajo manteniendo los formatos de la fila de encabezado original
    if max_segmentos > 1:
        ws.insert_rows(fila_encabezado + 1, max_segmentos - 1)
        # Copiar formato desde fila encabezado a filas insertadas
        for i in range(1, max_segmentos):
            copiar_fila_formato(ws, fila_encabezado, fila_encabezado + i, max_col)

    # Escribir las frases en celdas, una frase por fila y columna
    for col in range(1, max_col + 1):
        segmentos = titulos_segmentados[col - 1]
        for i, texto in enumerate(segmentos):
            fila = fila_encabezado + i
            celda = ws.cell(row=fila, column=col)
            celda.value = texto.strip()
            celda.font = copy(ws.cell(row=fila_encabezado, column=col).font)
            celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            celda.border = copy(ws.cell(row=fila_encabezado, column=col).border)

    # Combinar las celdas para frases repetidas y contiguas en filas superiores hasta la penúltima fila
    for nivel in range(max_segmentos - 1):
        texto_anterior = None
        inicio_col = None
        for col in range(1, max_col + 2):  # +1 para manejar el límite
            if col <= max_col:
                celda_actual = ws.cell(row=fila_encabezado + nivel, column=col)
                texto_actual = celda_actual.value
            else:
                texto_actual = None  # Fuerza cierre de grupo

            if texto_actual == texto_anterior and texto_actual is not None:
                # Continuar grupo
                pass
            else:
                # Cerrar grupo si hay más de 1 celda consecutiva
                if texto_anterior is not None and inicio_col is not None and (col - inicio_col) > 1:
                    rango_inicio = get_column_letter(inicio_col)
                    rango_fin = get_column_letter(col - 1)
                    rango_combinado = f"{rango_inicio}{fila_encabezado + nivel}:{rango_fin}{fila_encabezado + nivel}"
                    ws.merge_cells(rango_combinado)
                    celda_merge = ws.cell(row=fila_encabezado + nivel, column=inicio_col)
                    celda_merge.alignment = Alignment(horizontal="center", vertical="center")
                texto_anterior = texto_actual
                inicio_col = col

    return fila_encabezado + max_segmentos - 1


def reemplazar_encabezados_xlsx(xlsx_path, json_path=None):
    reemplazos_encabezados, reemplazos_de_nombres = cargar_config_json(json_path)

    wb = load_workbook(xlsx_path)
    ws = wb.active

    base_name = os.path.splitext(os.path.basename(xlsx_path))[0]
    nombre_celda_a1 = reemplazos_de_nombres.get(base_name, "Planilla")

    # Escribir nombre planilla en A1
    celda_a1 = ws.cell(row=1, column=1)
    celda_a1.value = nombre_celda_a1
    celda_a1.font = Font(name="Calibri", size=14, bold=True)
    celda_a1.alignment = Alignment(horizontal="left", vertical="center")

    fila_encabezado = 6

    # Procesar títulos divididos y combos, y obtener fila última ocupada del encabezado
    ultima_fila_encabezado = procesar_encabezados_divididos(ws, fila_encabezado, reemplazos_encabezados)

    color_azul = "0C769E"
    color_gris = "808080"
    font_fecha = Font(name="Calibri", size=11)
    alineacion_izq = Alignment(horizontal="left", vertical="center")

    # Escribir fecha en A2
    celda_a2 = ws.cell(row=2, column=1)
    fecha_hora_str = datetime.now().strftime("%d-%m-%y %H:%M")
    celda_a2.value = f"Fecha de generación: {fecha_hora_str}"
    celda_a2.font = font_fecha
    celda_a2.alignment = alineacion_izq

    # Rellenos gris y azul en I2, I3
    celda_i2 = ws.cell(row=2, column=9)
    celda_i2.fill = PatternFill("solid", fgColor=color_gris)
    celda_i3 = ws.cell(row=3, column=9)
    celda_i3.fill = PatternFill("solid", fgColor=color_azul)

    # Textos en J2 y J3
    celda_j2 = ws.cell(row=2, column=10)
    celda_j2.value = "Columnas solo lectura"
    celda_j2.alignment = alineacion_izq
    celda_j3 = ws.cell(row=3, column=10)
    celda_j3.value = "Columnas modificables"
    celda_j3.alignment = alineacion_izq

    # Contar filas con datos debajo de la última fila de encabezados
    fila_inicio_datos = ultima_fila_encabezado + 1
    total_filas = 0
    for row in ws.iter_rows(min_row=fila_inicio_datos, max_row=ws.max_row):
        if any(cell.value not in (None, "") for cell in row):
            total_filas += 1

    # Escribir total activos en A3
    celda_a3 = ws.cell(row=3, column=1)
    celda_a3.value = f"Total Activos: {total_filas}"
    celda_a3.font = Font(name="Calibri", size=11)
    celda_a3.alignment = Alignment(horizontal="left", vertical="center")

    wb.save(xlsx_path)
    print(f"Reemplazo de encabezados y anotaciones completo en {xlsx_path}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python script_3.py <archivo.xlsx>")
        sys.exit(1)

    archivo_xlsx = sys.argv[1]
    reemplazar_encabezados_xlsx(archivo_xlsx)
