import csv
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def es_celda_merged(ws, row, col):
    for merged in ws.merged_cells.ranges:
        if (row, col) in merged:
            if merged.min_row == row and merged.min_col == col:
                return False
            return True
    return False

def convertir_a_string(valor):
    if isinstance(valor, tuple):
        return ' '.join(convertir_a_string(v) for v in valor)
    if valor is None:
        return ""
    try:
        return str(valor)
    except Exception:
        return ""

def crear_excel_con_combinacion(csv_path, xlsx_path):
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Registros"
        
        # Encabezados para formato especial según modelo
        encabezados_gris = [
            "val-sql", "val-activos", "id", "nombre", "estado", "Nodo",
            "Propietario", "Decreto", "Categoría VU", "Estructuras", "OOCC",
            "Armario", "DEA", "Línea", "Subestación", "Comuna", "Patio",
            "Antecedentes complementarios", "Sala SSGG y/o Caseta", "Tramo",
            "Nodo 1", "Nodo 2", "Nodo 3", "Circuito", "SS.GG., Caseta y Salas de ER",
            "Sistema Eléctrico", "Nodo Paño/Barra", "Nodo SSEE", "Paño", "Servidumbres",
            "Vano", "Nodo SS.GG. y Caseta", "Nodo Paño", "Nodo de Estructura", 
            "inf-id"
        ]
        
        color_gris = "808080"
        color_azul = "0C769E"
        font_encabezado = Font(bold=True, color="FFFFFF")
        align_centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
        borde_delgado = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        
        with open(csv_path, newline='', encoding='utf-8') as f:
            reader = csv.reader(f, delimiter=';')
            filas = [row for row in reader if row]
        
        n_encabezado_filas = 5  # dejar filas en blanco arriba según requerimiento anterior
        n_cols = max(len(row) for row in filas) if filas else 0
        
        # Aplicar encabezados en filas 1 a 5 (vacías o personalizables)
        # En este ejemplo no se llenan para dejar vacío pero ya se reservan.
        
        # Insertar datos desde fila 6
        for r_idx, fila in enumerate(filas, start=n_encabezado_filas + 1):
            for c_idx, valor in enumerate(fila, start=1):
                celda = ws.cell(row=r_idx, column=c_idx)
                valor_str = convertir_a_string(valor)
                celda.value = valor_str
                
                # Formato para encabezado (fila 6) basado en lista encabezados
                if r_idx == n_encabezado_filas + 1:
                    if valor_str in encabezados_gris:
                        celda.fill = PatternFill("solid", fgColor=color_gris)
                    else:
                        celda.fill = PatternFill("solid", fgColor=color_azul)
                    celda.font = font_encabezado
                    celda.alignment = align_centro
                    celda.border = borde_delgado
                else:
                    # Formato para datos normales
                    celda.alignment = Alignment(horizontal="left", vertical="center")
                    celda.border = borde_delgado
        
        # Ajustar ancho columnas
        for col in range(1, n_cols + 1):
            letra_col = ws.cell(row=n_encabezado_filas + 1, column=col).column_letter
            ws.column_dimensions[letra_col].width = 15
        
        wb.save(xlsx_path)
        
        # Borrar CSV después de crear XLSX
        if os.path.exists(csv_path):
            os.remove(csv_path)
        
        return "Listo"
    except Exception as e:
        return f"Error: {str(e)}"

if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("Uso: python Planillas_xlsx.py <archivo.csv>")
        sys.exit(1)
    csv_path = sys.argv[1]
    base_name = os.path.splitext(csv_path)[0]
    xlsx_path = base_name + ".xlsx"
    resultado = crear_excel_con_combinacion(csv_path, xlsx_path)
    print(resultado)


