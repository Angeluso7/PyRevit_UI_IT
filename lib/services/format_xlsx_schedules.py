# -*- coding: utf-8 -*-
"""
lib/services/format_xlsx_schedules.py  (CPython)
Post-procesa un XLSX generado por csv_to_xlsx:
  - A1: nombre legible de la planilla (desde script.json)
  - A2: fecha de generación
  - A3: total de activos
  - I2/I3: rellenos de referencia  J2/J3: leyenda
  - Fila 6: encabezados divididos por '/' con merge de celdas contiguas
Uso: python format_xlsx_schedules.py <ruta_xlsx>
"""
import sys
import os
import json
from copy import copy
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ── Ruta a script.json ────────────────────────────────────────────────────────
_BASE_DATA = os.path.join(
    os.path.expanduser('~'),
    r'AppData\Roaming\MyPyRevitExtention\PyRevitIT.extension\data'
)
_SCRIPT_JSON = os.path.join(_BASE_DATA, 'script.json')

FILA_ENCABEZADO = 6


def cargar_config(json_path=None):
    path = json_path or _SCRIPT_JSON
    if not os.path.exists(path):
        return {}, {}
    with open(path, 'r', encoding='utf-8') as f:
        cfg = json.load(f)
    return cfg.get('reemplazos_encabezados', {}), cfg.get('reemplazos_de_nombres', {})


def _copiar_formato(ws, fila_src, fila_dst, max_col):
    for col in range(1, max_col + 1):
        src = ws.cell(row=fila_src, column=col)
        dst = ws.cell(row=fila_dst, column=col)
        dst.font       = copy(src.font)
        dst.border     = copy(src.border)
        dst.fill       = copy(src.fill)
        dst.alignment  = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)


def procesar_encabezados(ws, reemplazos):
    max_col = ws.max_column
    titulos = []
    max_seg = 1

    for col in range(1, max_col + 1):
        val = ws.cell(row=FILA_ENCABEZADO, column=col).value or ''
        val = reemplazos.get(val, val)
        segs = val.split('/')
        titulos.append(segs)
        if len(segs) > max_seg:
            max_seg = len(segs)

    if max_seg > 1:
        ws.insert_rows(FILA_ENCABEZADO + 1, max_seg - 1)
        for i in range(1, max_seg):
            _copiar_formato(ws, FILA_ENCABEZADO, FILA_ENCABEZADO + i, max_col)

    for col, segs in enumerate(titulos, start=1):
        for i, texto in enumerate(segs):
            cell = ws.cell(row=FILA_ENCABEZADO + i, column=col)
            cell.value     = texto.strip()
            cell.alignment = Alignment(horizontal='center', vertical='center',
                                       wrap_text=True)

    # Merge de celdas contiguas con mismo texto en niveles superiores
    for nivel in range(max_seg - 1):
        prev_txt, col_ini = None, None
        for col in range(1, max_col + 2):
            txt = (ws.cell(row=FILA_ENCABEZADO + nivel, column=col).value
                   if col <= max_col else None)
            if txt == prev_txt and txt is not None:
                pass
            else:
                if prev_txt is not None and col_ini is not None and (col - col_ini) > 1:
                    rng = '{}{0}:{}{1}'.format(
                        get_column_letter(col_ini), FILA_ENCABEZADO + nivel,
                        get_column_letter(col - 1), FILA_ENCABEZADO + nivel
                    )
                    ws.merge_cells(rng)
                    ws.cell(row=FILA_ENCABEZADO + nivel,
                            column=col_ini).alignment = Alignment(
                        horizontal='center', vertical='center')
                prev_txt, col_ini = txt, col

    return FILA_ENCABEZADO + max_seg - 1


def anotar_cabecera(ws, nombre_planilla, ultima_fila_enc):
    font_titulo = Font(name='Calibri', size=14, bold=True)
    font_normal = Font(name='Calibri', size=11)
    align_izq   = Alignment(horizontal='left', vertical='center')

    # A1 — nombre
    c = ws.cell(row=1, column=1)
    c.value, c.font, c.alignment = nombre_planilla, font_titulo, align_izq

    # A2 — fecha
    c = ws.cell(row=2, column=1)
    c.value     = 'Fecha de generación: {}'.format(
        datetime.now().strftime('%d-%m-%y %H:%M'))
    c.font, c.alignment = font_normal, align_izq

    # I2 gris / I3 azul
    ws.cell(row=2, column=9).fill = PatternFill('solid', fgColor='808080')
    ws.cell(row=3, column=9).fill = PatternFill('solid', fgColor='0C769E')

    # J2 / J3 leyenda
    for row, txt in [(2, 'Columnas solo lectura'), (3, 'Columnas modificables')]:
        c = ws.cell(row=row, column=10)
        c.value, c.alignment = txt, align_izq

    # A3 — total activos
    fila_datos = ultima_fila_enc + 1
    total = sum(
        1 for row in ws.iter_rows(min_row=fila_datos, max_row=ws.max_row)
        if any(cell.value not in (None, '') for cell in row)
    )
    c = ws.cell(row=3, column=1)
    c.value     = 'Total Activos: {}'.format(total)
    c.font, c.alignment = font_normal, align_izq


def procesar(xlsx_path, json_path=None):
    reemplazos_enc, reemplazos_nom = cargar_config(json_path)

    wb = load_workbook(xlsx_path)
    ws = wb.active

    base          = os.path.splitext(os.path.basename(xlsx_path))[0]
    nombre_celda  = reemplazos_nom.get(base, base)

    ultima_fila   = procesar_encabezados(ws, reemplazos_enc)
    anotar_cabecera(ws, nombre_celda, ultima_fila)

    wb.save(xlsx_path)
    print('OK: {}'.format(xlsx_path))


def main():
    if len(sys.argv) < 2:
        print('Uso: format_xlsx_schedules.py <ruta_xlsx>')
        sys.exit(1)

    xlsx_path = sys.argv[1]
    if not os.path.exists(xlsx_path):
        print('Error: no existe: {}'.format(xlsx_path))
        sys.exit(1)

    json_path = sys.argv[2] if len(sys.argv) > 2 else None

    try:
        procesar(xlsx_path, json_path)
    except Exception as e:
        print('Error: {}'.format(e))
        sys.exit(1)


if __name__ == '__main__':
    main()