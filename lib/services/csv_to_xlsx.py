# -*- coding: utf-8 -*-
"""
lib/services/csv_to_xlsx.py  (CPython)
Convierte un CSV (delimitador ;) a XLSX con formato base.
Uso: python csv_to_xlsx.py <ruta_csv>
El XLSX se crea en la misma carpeta con el mismo nombre base.
Al finalizar con éxito borra el CSV.
"""
import sys
import os
import csv

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ── Columnas que van en gris (solo lectura) ───────────────────────────────────
COLS_GRIS = {
    "val-sql", "val-activos", "id", "nombre", "estado", "Nodo",
    "Propietario", "Decreto", "Categoría VU", "Estructuras", "OOCC",
    "Armario", "DEA", "Línea", "Subestación", "Comuna", "Patio",
    "Antecedentes complementarios", "Sala SSGG y/o Caseta", "Tramo",
    "Nodo 1", "Nodo 2", "Nodo 3", "Circuito",
    "SS.GG., Caseta y Salas de ER", "Sistema Eléctrico",
    "Nodo Paño/Barra", "Nodo SSEE", "Paño", "Servidumbres",
    "Vano", "Nodo SS.GG. y Caseta", "Nodo Paño",
    "Nodo de Estructura", "inf-id",
}

COLOR_GRIS  = "808080"
COLOR_AZUL  = "0C769E"
N_FILAS_CABECERA = 5   # filas 1-5 reservadas (A1..A3 las escribe format_xlsx)


def _borde_delgado():
    s = Side(style='thin')
    return Border(left=s, right=s, top=s, bottom=s)


def crear_excel(csv_path, xlsx_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Registros"

    font_enc    = Font(bold=True, color="FFFFFF")
    align_enc   = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)
    align_datos = Alignment(horizontal="left", vertical="center")
    borde       = _borde_delgado()

    with open(csv_path, newline='', encoding='utf-8') as f:
        filas = [row for row in csv.reader(f, delimiter=';') if row]

    n_cols = max((len(r) for r in filas), default=0)

    for r_idx, fila in enumerate(filas, start=N_FILAS_CABECERA + 1):
        for c_idx, valor in enumerate(fila, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=valor or "")
            cell.border = borde
            if r_idx == N_FILAS_CABECERA + 1:          # fila de encabezado
                cell.font  = font_enc
                cell.alignment = align_enc
                cell.fill  = PatternFill(
                    "solid",
                    fgColor=COLOR_GRIS if valor in COLS_GRIS else COLOR_AZUL
                )
            else:
                cell.alignment = align_datos

    # Ancho fijo 15 para todas las columnas
    for col in range(1, n_cols + 1):
        letra = ws.cell(row=N_FILAS_CABECERA + 1, column=col).column_letter
        ws.column_dimensions[letra].width = 15

    wb.save(xlsx_path)

    # Borrar CSV
    try:
        os.remove(csv_path)
    except OSError:
        pass


def main():
    if len(sys.argv) < 2:
        print("Uso: csv_to_xlsx.py <ruta_csv>")
        sys.exit(1)

    csv_path  = sys.argv[1]
    xlsx_path = os.path.splitext(csv_path)[0] + '.xlsx'

    if not os.path.exists(csv_path):
        print("Error: no existe CSV: {}".format(csv_path))
        sys.exit(1)

    try:
        crear_excel(csv_path, xlsx_path)
        print("OK: {}".format(xlsx_path))
    except Exception as e:
        print("Error: {}".format(e))
        sys.exit(1)


if __name__ == '__main__':
    main()