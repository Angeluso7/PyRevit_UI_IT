# -*- coding: utf-8 -*-
"""
formatear_tablas_excel_v2.py

Formato VERTICAL con colores por parámetro:
- Columna A: nombres de parámetros (filas)
- Columnas B, C, D...: valores por elemento
- Colores desde colores_parametros.json

Reglas de color:
1. Parámetros en JSON → usar color del JSON
2. Parámetros NO en JSON → FFFFFFCC (amarillo claro)
3. Parámetros adicionales (ElementId, Categoría, etc.) → D7D7D7 (gris medio)
4. CodIntBIM → D7D7D7 (gris medio, ignorando JSON)
"""

import sys
import os
import json
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Ruta de colores
DATA_DIR = os.path.join(
    os.path.expanduser("~"),
    r"AppData\Roaming\MyPyRevitExtention\PyRevitIT.extension\data"
)
COLORES_JSON_PATH = os.path.join(DATA_DIR, "colores_parametros.json")

# Colores por defecto
COLOR_DEFAULT_NO_JSON = "FFFFFFCC"    # amarillo claro para parámetros sin color en JSON
COLOR_ADICIONALES = "FFD7D7D7"        # gris medio para parámetros adicionales
COLOR_CODINTBIM = "FFD7D7D7"          # gris medio para CodIntBIM (destacado)
COLOR_HEADER_INDICE = "FF366092"      # azul para índice/excepciones
FONT_COLOR_INDICE = "FFFFFFFF"
BORDER_STYLE = Side(style="thin", color="FF000000")

def cargar_json(ruta):
    with open(ruta, "r", encoding="utf-8") as f:
        return json.load(f)

def cargar_colores_parametros():
    """Carga el JSON de colores. Si falla, devuelve dict vacío."""
    if not os.path.exists(COLORES_JSON_PATH):
        print("AVISO: no se encontró colores_parametros.json")
        return {}
    try:
        with open(COLORES_JSON_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
        # Limpiar espacios en los valores de color
        return {k.strip(): v.strip() for k, v in data.items()}
    except Exception as e:
        print("Error cargando colores: {}".format(e))
        return {}

def aplicar_estilo_celda(cell, color_hex, bold=False, fuente_color="FF000000"):
    cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
    cell.font = Font(bold=bold, color=fuente_color)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = Border(
        left=BORDER_STYLE,
        right=BORDER_STYLE,
        top=BORDER_STYLE,
        bottom=BORDER_STYLE,
    )

def autoajustar_columnas(ws):
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row_idx in range(1, ws.max_row + 1):
            v = ws.cell(row_idx, col_idx).value
            if v:
                max_len = max(max_len, len(str(v)))
        ancho = min(max(max_len + 2, 15), 60)
        ws.column_dimensions[col_letter].width = ancho

def autoajustar_filas(ws):
    for row_idx in range(1, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 20

def generar_hoja_indice(ws, listado_tablas):
    ws.title = "Índice"

    ws["B3"] = "Listado de Tablas"
    ws["B3"].font = Font(bold=True, size=12)

    ws["B4"] = "Código"
    ws["C4"] = "Nombre Tabla (Schedule)"
    for col in ("B", "C"):
        aplicar_estilo_celda(ws[col + "4"], COLOR_HEADER_INDICE, bold=True, fuente_color=FONT_COLOR_INDICE)

    valores = listado_tablas.get("valores", [])
    claves = listado_tablas.get("claves", [])

    pares = list(zip(valores, claves))
    pares.sort(key=lambda x: x[0])

    for idx, (codigo, nombre_schedule) in enumerate(pares):
        fila = 5 + idx
        ws["B{}".format(fila)] = codigo
        ws["C{}".format(fila)] = nombre_schedule

    autoajustar_columnas(ws)

def generar_hoja_excepciones(ws, excepciones):
    ws.title = "Excepciones"

    ws["A1"] = "Excepciones durante la extracción"
    ws["A1"].font = Font(bold=True, size=12)

    campos = ["ElementId", "CodIntBIM", "Categoría", "Familia", "Tipo", "Nombre_RVT", "Situación"]

    for col_idx, campo in enumerate(campos, 1):
        ws.cell(1, col_idx, value=campo)
        aplicar_estilo_celda(ws.cell(1, col_idx), COLOR_HEADER_INDICE, bold=True, fuente_color=FONT_COLOR_INDICE)

    for row_idx, exc in enumerate(excepciones, 2):
        elem = exc.get("elemento", {}) or {}
        situacion = exc.get("situacion", "")

        ws.cell(row_idx, 1, value=str(elem.get("ElementId", "")))
        ws.cell(row_idx, 2, value=elem.get("CodIntBIM", ""))
        ws.cell(row_idx, 3, value=elem.get("Categoría", ""))
        ws.cell(row_idx, 4, value=elem.get("Familia", ""))
        ws.cell(row_idx, 5, value=elem.get("Tipo", ""))
        ws.cell(row_idx, 6, value=elem.get("Nombre_RVT", ""))
        ws.cell(row_idx, 7, value=situacion)

        for col_idx in range(1, 8):
            cell = ws.cell(row_idx, col_idx)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = Border(
                left=BORDER_STYLE,
                right=BORDER_STYLE,
                top=BORDER_STYLE,
                bottom=BORDER_STYLE,
            )

    autoajustar_columnas(ws)

def obtener_color_parametro(param, colores_dict):
    """
    Determina el color según reglas:
    1. CodIntBIM → gris medio (COLOR_CODINTBIM)
    2. Parámetros adicionales → gris medio (COLOR_ADICIONALES)
    3. Si está en JSON → usar color del JSON
    4. Si NO está en JSON → amarillo claro (COLOR_DEFAULT_NO_JSON)
    """
    # Lista de parámetros adicionales que incorporamos nosotros
    parametros_adicionales = [
        "ElementId",
        "Categoría",
        "Familia",
        "Tipo",
        "Nombre_RVT",
        "Situación CodIntBIM",
        "Elementos",
    ]
    
    # Regla 1: CodIntBIM siempre gris medio
    if param == "CodIntBIM":
        return COLOR_CODINTBIM
    
    # Regla 2: Parámetros adicionales → gris medio
    if param in parametros_adicionales:
        return COLOR_ADICIONALES
    
    # Regla 3: Si está en JSON, usar ese color
    if param in colores_dict:
        return colores_dict[param]
    
    # Regla 4: Si NO está en JSON → amarillo claro
    return COLOR_DEFAULT_NO_JSON

def generar_hoja_cm_vertical(ws, codigo_cm, elementos, headers_schedule, colores_dict):
    """
    Formato vertical:
    - Fila por parámetro (columna A)
    - Columnas por elemento (B, C, D...)
    Orden:
    1) CodIntBIM
    2) headers_schedule (tabla CM correspondiente)
    3) parámetros extra fijos
    """
    if not elementos:
        ws["A1"] = "Sin datos"
        return

    # Parámetros extra fijos (al final)
    campos_extra = [
        "ElementId",
        "Categoría",
        "Familia",
        "Tipo",
        "Nombre_RVT",
        "Situación CodIntBIM",
        "Elementos",
    ]

    # Headers base desde schedule, sin duplicar extras ni CodIntBIM
    base_headers = [
        h for h in (headers_schedule or [])
        if h and h not in campos_extra and h != "CodIntBIM"
    ]

    # Orden final
    parametros_orden = []
    parametros_orden.append("CodIntBIM")  # 1) siempre primero

    for h in base_headers:                # 2) headers de planilla
        if h not in parametros_orden:
            parametros_orden.append(h)

    for h in campos_extra:                # 3) extras
        if h not in parametros_orden:
            parametros_orden.append(h)

    # Conteo por CodIntBIM para "único"/"varios"
    conteo_por_codint = defaultdict(int)
    for elem in elementos:
        c = (elem.get("CodIntBIM") or "").strip()
        if c:
            conteo_por_codint[c] += 1

    # Columna A: nombre de parámetros + colores
    for row_idx, param in enumerate(parametros_orden, 1):
        cell = ws.cell(row=row_idx, column=1)
        cell.value = param
        
        # Obtener color según reglas
        color = obtener_color_parametro(param, colores_dict)
        aplicar_estilo_celda(cell, color, bold=True)

    # Columnas B...: valores por elemento
    for col_idx, elem in enumerate(elementos, 2):
        codint = (elem.get("CodIntBIM") or "").strip()
        mult = conteo_por_codint.get(codint, 0)
        val_elementos = "varios" if mult > 1 else "único"

        for row_idx, param in enumerate(parametros_orden, 1):
            cell = ws.cell(row=row_idx, column=col_idx)

            if param == "ElementId":
                valor = str(elem.get("ElementId", ""))
            elif param == "CodIntBIM":
                valor = elem.get("CodIntBIM", "")
            elif param == "Categoría":
                valor = elem.get("Categoría", "")
            elif param == "Familia":
                valor = elem.get("Familia", "")
            elif param == "Tipo":
                valor = elem.get("Tipo", "")
            elif param == "Nombre_RVT":
                valor = elem.get("Nombre_RVT", "")
            elif param == "Situación CodIntBIM":
                valor = "Elemento no Anidado"
            elif param == "Elementos":
                valor = val_elementos
            else:
                valor = elem.get(param, "")

            cell.value = valor if valor else ""
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = Border(
                left=BORDER_STYLE,
                right=BORDER_STYLE,
                top=BORDER_STYLE,
                bottom=BORDER_STYLE,
            )

    ws.freeze_panes = "B1"
    autoajustar_columnas(ws)
    autoajustar_filas(ws)

def main(ruta_json, ruta_xlsx_salida):
    print("=" * 70)
    print("GENERACIÓN DE EXCEL - FORMATO VERTICAL CON COLORES")
    print("=" * 70)

    datos = cargar_json(ruta_json)

    elementos_por_tabla = datos.get("elementos_por_tabla", {})
    listado_tablas = datos.get("listado_tablas", {})
    headers_por_tabla = datos.get("headers_por_tabla", {})
    excepciones = datos.get("excepciones", [])

    print(" - Tablas CM: {}".format(len(elementos_por_tabla)))

    colores_dict = cargar_colores_parametros()
    print(" - Colores cargados: {}".format(len(colores_dict)))

    wb = Workbook()
    if wb.sheetnames:
        wb.remove(wb.active)

    # Índice
    ws_idx = wb.create_sheet("Índice", 0)
    generar_hoja_indice(ws_idx, listado_tablas)

    # Hojas CM (una por código)
    codigos_ordenados = sorted(elementos_por_tabla.keys())
    for i, codigo_cm in enumerate(codigos_ordenados, start=1):
        elementos = elementos_por_tabla.get(codigo_cm, [])
        headers_schedule = headers_por_tabla.get(codigo_cm, [])
        ws_cm = wb.create_sheet(codigo_cm[:31], i)
        generar_hoja_cm_vertical(ws_cm, codigo_cm, elementos, headers_schedule, colores_dict)

    # Excepciones
    if excepciones:
        ws_exc = wb.create_sheet("Excepciones", len(wb.sheetnames))
        generar_hoja_excepciones(ws_exc, excepciones)

    wb.save(ruta_xlsx_salida)
    print("✓ Archivo generado: {}".format(ruta_xlsx_salida))
    print("=" * 70)

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Uso: formatear_tablas_excel_v2.py <ruta_json> <ruta_xlsx_salida>")
        sys.exit(1)
    main(sys.argv[1], sys.argv[2])