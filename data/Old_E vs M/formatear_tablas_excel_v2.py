# -*- coding: utf-8 -*-
"""
Script de formateo de Excel v2
Genera Excel con estructura de tablas, hoja índice y hoja de excepciones
"""

import sys
import json
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_COLOR = 'FF366092'        # azul encabezado base
HEADER_FONT_COLOR = 'FFFFFFFF'
BORDER_STYLE = Side(style='thin', color='FF000000')
HEADER_NARANJA = 'FFF4B084'      # naranja para columnas desde CodIntBIM


def cargar_datos_json(ruta_json):
    try:
        with open(ruta_json, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print("Error al leer JSON: {}".format(e), file=sys.stderr)
        raise SystemExit


def aplicar_formato_encabezado(ws, row_num, headers):
    """Encabezado azul general + recolor naranja desde CodIntBIM a la derecha."""
    ncols = len(headers)
    header_font = Font(bold=True, color=HEADER_FONT_COLOR)
    fill_azul = PatternFill(start_color=HEADER_COLOR,
                            end_color=HEADER_COLOR,
                            fill_type='solid')
    fill_naranja = PatternFill(start_color=HEADER_NARANJA,
                               end_color=HEADER_NARANJA,
                               fill_type='solid')

    # Índice de CodIntBIM (si existe)
    idx_cod = None
    for i, h in enumerate(headers):
        if (h or "").strip() == "CodIntBIM":
            idx_cod = i
            break

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col_idx)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center',
                                   vertical='center',
                                   wrap_text=True)
        cell.border = Border(left=BORDER_STYLE, right=BORDER_STYLE,
                             top=BORDER_STYLE, bottom=BORDER_STYLE)

        # Color base azul
        cell.fill = fill_azul

        # Desde CodIntBIM a la derecha -> naranja
        if idx_cod is not None and col_idx - 1 >= idx_cod:
            cell.fill = fill_naranja


def auto_ajustar_columnas(ws):
    for col in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col)
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)


def generar_hoja_indice(ws, listado_tablas):
    ws.title = "Índice"
    ws['B3'] = "Listado de Tablas"
    ws['B3'].font = Font(bold=True, size=12)

    headers = ["Código", "Nombre Tabla (Schedule)"]
    ws['B4'] = headers[0]
    ws['C4'] = headers[1]
    aplicar_formato_encabezado(ws, 4, headers)

    valores = listado_tablas.get('valores', [])
    claves = listado_tablas.get('claves', [])

    pares = list(zip(valores, claves))
    pares.sort(key=lambda x: x[0])

    for idx, (codigo, nombre_schedule) in enumerate(pares):
        fila = 5 + idx
        ws['B{}'.format(fila)] = codigo
        ws['C{}'.format(fila)] = nombre_schedule

    auto_ajustar_columnas(ws)


def generar_hoja_tabla(ws, codigo_tabla, elementos, headers_por_tabla):
    if not elementos:
        return

    # Campos extra fijos que siempre queremos al final
    campos_extra = [
        'ElementId', 'CodIntBIM', 'Categoría', 'Familia', 'Tipo',
        'Nombre_RVT', 'Situación CodIntBIM', 'Elementos'
    ]

    # Headers base desde ViewSchedule (desde val-sql a la derecha, en orden)
    # Estos serán los ÚNICOS parámetros que se usarán para datos,
    # es decir, no se añadirán parámetros adicionales que no estén aquí.
    base_headers = headers_por_tabla.get(codigo_tabla, []) or []
    base_headers = [h for h in base_headers if h]  # quitar vacíos

    # Evitar duplicar campos extra dentro de base_headers
    headers_params = [h for h in base_headers if h not in campos_extra]

    # Orden final de columnas: primero los parámetros del schedule,
    # luego los campos extra de control.
    headers_finales = headers_params + campos_extra

    # Escribir encabezados
    for col_idx, h in enumerate(headers_finales, 1):
        ws.cell(row=1, column=col_idx, value=h)
    aplicar_formato_encabezado(ws, 1, headers_finales)

    # Conteo por CodIntBIM para "único" / "varios"
    conteo_por_codint = defaultdict(int)
    for elem in elementos:
        c = (elem.get('CodIntBIM') or '').strip()
        if c:
            conteo_por_codint[c] += 1

    # Escribir datos: SOLO se tomarán los parámetros cuyos nombres
    # están en headers_params; si el elemento tiene más parámetros,
    # se ignoran en la exportación.
    for row_idx, elem in enumerate(elementos, 2):
        codint = (elem.get('CodIntBIM') or '').strip()
        mult = conteo_por_codint.get(codint, 0)
        val_elementos = "varios" if mult > 1 else "único"

        for col_idx, h in enumerate(headers_finales, 1):
            if h == 'ElementId':
                v = str(elem.get('ElementId', ''))
            elif h == 'CodIntBIM':
                v = elem.get('CodIntBIM', '')
            elif h == 'Categoría':
                v = elem.get('Categoría', '')
            elif h == 'Familia':
                v = elem.get('Familia', '')
            elif h == 'Tipo':
                v = elem.get('Tipo', '')
            elif h == 'Nombre_RVT':
                v = elem.get('Nombre_RVT', '')
            elif h == 'Situación CodIntBIM':
                # En el futuro se puede leer un flag de anidado desde el JSON
                v = "Elemento no Anidado"
            elif h == 'Elementos':
                v = val_elementos
            else:
                # Parámetro normal: se usa SOLO si está en headers_params
                v = elem.get(h, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=v)
            cell.alignment = Alignment(horizontal='left',
                                       vertical='top',
                                       wrap_text=True)
            cell.border = Border(left=BORDER_STYLE, right=BORDER_STYLE,
                                 top=BORDER_STYLE, bottom=BORDER_STYLE)

    ws.freeze_panes = 'A2'
    auto_ajustar_columnas(ws)


def generar_hoja_excepciones(ws, excepciones):
    if not excepciones:
        return

    campos = [
        'ElementId', 'CodIntBIM', 'Categoría', 'Familia', 'Tipo',
        'Nombre_RVT', 'Situación CodIntBIM', 'Elementos'
    ]

    for col_idx, h in enumerate(campos, 1):
        ws.cell(row=1, column=col_idx, value=h)
    aplicar_formato_encabezado(ws, 1, campos)

    for row_idx, exc in enumerate(excepciones, 2):
        elem = exc.get('elemento', {}) or {}
        situacion = exc.get('situacion', 'No existe')

        for col_idx, h in enumerate(campos, 1):
            if h == 'ElementId':
                v = str(elem.get('ElementId', ''))
            elif h == 'CodIntBIM':
                v = elem.get('CodIntBIM', situacion)
            elif h == 'Categoría':
                v = elem.get('Categoría', '')
            elif h == 'Familia':
                v = elem.get('Familia', '')
            elif h == 'Tipo':
                v = elem.get('Tipo', '')
            elif h == 'Nombre_RVT':
                v = elem.get('Nombre_RVT', '')
            elif h == 'Situación CodIntBIM':
                v = "Elemento no Anidado"
            elif h == 'Elementos':
                v = "único"
            else:
                v = ""
            cell = ws.cell(row=row_idx, column=col_idx, value=v)
            cell.alignment = Alignment(horizontal='left',
                                       vertical='top',
                                       wrap_text=True)
            cell.border = Border(left=BORDER_STYLE, right=BORDER_STYLE,
                                 top=BORDER_STYLE, bottom=BORDER_STYLE)

    ws.freeze_panes = 'A2'
    auto_ajustar_columnas(ws)


def main(ruta_json, ruta_xlsx_salida):
    datos = cargar_datos_json(ruta_json)
    elementos_por_tabla = datos.get('elementos_por_tabla', {})
    excepciones = datos.get('excepciones', [])
    listado_tablas = datos.get('listado_tablas', {})
    headers_por_tabla = datos.get('headers_por_tabla', {})

    wb = Workbook()
    if wb.sheetnames:
        wb.remove(wb.active)

    ws_idx = wb.create_sheet('Índice', 0)
    generar_hoja_indice(ws_idx, listado_tablas)

    codigos_ordenados = sorted(elementos_por_tabla.keys())
    sheet_idx = 1
    for codigo in codigos_ordenados:
        elems = elementos_por_tabla.get(codigo, [])
        ws = wb.create_sheet(codigo[:31], sheet_idx)
        generar_hoja_tabla(ws, codigo, elems, headers_por_tabla)
        sheet_idx += 1

    if excepciones:
        ws_exc = wb.create_sheet('Excepciones', sheet_idx)
        generar_hoja_excepciones(ws_exc, excepciones)

    wb.save(ruta_xlsx_salida)
    print("Archivo generado: {}".format(ruta_xlsx_salida))


if __name__ == '__main__':
    if len(sys.argv) < 3:
        print("Uso: formatear_tablas_excel_v2.py <ruta_json> <ruta_xlsx_salida>",
              file=sys.stderr)
        sys.exit(1)
    main(sys.argv[1], sys.argv[2])
