# -*- coding: utf-8 -*-
"""
formatear_tablas_excel.py
Lee comparacion.json y genera:
  - Hoja Índice
  - Hojas por código CMxx con Tabla comparativa
  - Colores por estado y leyenda en cada hoja
"""

import sys
import json
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

COLOR_ESTADOS = {
    'ok': 'FFC6EFCE',           # verde claro
    'falta_modelo': 'FFFFC7CE', # rojo claro
    'falta_excel': 'FFFFEB9C',  # amarillo
    'difiere': 'FFF4B084',      # naranja
    'no_existe': 'FF66FFFF',    # gris (no existe en ninguna fuente)
}

HEADER_COLOR = 'FF366092'
HEADER_FONT_COLOR = 'FFFFFFFF'
BORDER_STYLE = Side(style='thin', color='FF000000')


def cargar_datos_json(ruta_json):
    with open(ruta_json, 'r', encoding='utf-8') as f:
        return json.load(f)


def aplicar_formato_encabezado(ws, row_num, headers):
    header_font = Font(bold=True, color=HEADER_FONT_COLOR)
    header_fill = PatternFill(start_color=HEADER_COLOR,
                              end_color=HEADER_COLOR,
                              fill_type='solid')
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center',
                                   vertical='center',
                                   wrap_text=True)
        cell.border = Border(left=BORDER_STYLE, right=BORDER_STYLE,
                             top=BORDER_STYLE, bottom=BORDER_STYLE)


def auto_ajustar_columnas(ws):
    for col in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col)
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)


def agregar_leyenda(ws, fila_inicio):
    leyenda = [
        ('OK / Coincide', COLOR_ESTADOS['ok']),
        ('Falta en modelo', COLOR_ESTADOS['falta_modelo']),
        ('Falta en Excel', COLOR_ESTADOS['falta_excel']),
        ('Diferencia de valor', COLOR_ESTADOS['difiere']),
        ('No existe', COLOR_ESTADOS['no_existe']),
    ]
    fila = fila_inicio
    col = 1
    for texto, color in leyenda:
        cell_color = ws.cell(row=fila, column=col)
        cell_color.fill = PatternFill(start_color=color,
                                      end_color=color,
                                      fill_type='solid')
        cell_text = ws.cell(row=fila, column=col + 1)
        cell_text.value = texto
        fila += 1


def generar_hoja_indice(ws, listado_tablas):
    ws.title = "Índice"
    ws['B3'] = "Listado de Tablas"
    ws['B3'].font = Font(bold=True, size=12)

    headers = ["Código", "Nombre Tabla (Schedule)"]
    ws['B4'] = headers[0]
    ws['C4'] = headers[1]
    aplicar_formato_encabezado(ws, 4, headers)

    valores = listado_tablas.get('valores', [])
    claves = listado_tablas.get('claves', [])

    pares = list(zip(valores, claves))
    pares.sort(key=lambda x: x[0])

    for idx, (codigo, nombre_schedule) in enumerate(pares):
        fila = 5 + idx
        ws['B{}'.format(fila)] = codigo
        ws['C{}'.format(fila)] = nombre_schedule

    auto_ajustar_columnas(ws)


def escribir_tabla_comparativa(ws, headers, filas):
    if not filas:
        return 1

    fila_actual = 1

    # Encabezados
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=fila_actual, column=col_idx, value=h)
    aplicar_formato_encabezado(ws, fila_actual, headers)
    fila_actual += 1

    # Filas (ya fusionadas Modelo+Planilla)
    for fila in filas:
        valores = fila['valores']
        estados = fila['estado_por_celda']
        for col_idx, val in enumerate(valores, 1):
            cell = ws.cell(row=fila_actual, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal='left',
                                       vertical='top',
                                       wrap_text=True)
            cell.border = Border(left=BORDER_STYLE, right=BORDER_STYLE,
                                 top=BORDER_STYLE, bottom=BORDER_STYLE)
            estado = estados[col_idx - 1] if col_idx - 1 < len(estados) else 'ok'
            color = COLOR_ESTADOS.get(estado, COLOR_ESTADOS['ok'])
            cell.fill = PatternFill(start_color=color,
                                    end_color=color,
                                    fill_type='solid')
        fila_actual += 1

    return fila_actual


def generar_hoja_tabla(ws, bloque):
    headers = bloque.get('headers', [])
    filas = bloque.get('filas', [])

    fila_actual = escribir_tabla_comparativa(ws, headers, filas)

    fila_leyenda = fila_actual + 2
    agregar_leyenda(ws, fila_leyenda)

    ws.freeze_panes = 'A2'
    auto_ajustar_columnas(ws)


def main(ruta_json_comparacion, ruta_xlsx_salida):
    datos = cargar_datos_json(ruta_json_comparacion)

    listado_tablas = datos.get('listado_tablas', {})
    datos_por_tabla = datos.get('datos_por_tabla', {})

    wb = Workbook()
    if wb.sheetnames:
        wb.remove(wb.active)

    ws_idx = wb.create_sheet('Índice', 0)
    generar_hoja_indice(ws_idx, listado_tablas)

    codigos_ordenados = sorted(datos_por_tabla.keys())
    sheet_idx = 1
    for codigo in codigos_ordenados:
        bloque = datos_por_tabla.get(codigo, {})
        ws = wb.create_sheet(codigo[:31], sheet_idx)
        generar_hoja_tabla(ws, bloque)
        sheet_idx += 1

    wb.save(ruta_xlsx_salida)
    print("Archivo generado: {}".format(ruta_xlsx_salida))


if __name__ == '__main__':
    if len(sys.argv) < 3:
        print(
            "Uso: formatear_tablas_excel.py <ruta_json_comparacion> <ruta_xlsx_salida>",
            file=sys.stderr
        )
        sys.exit(1)

    ruta_json = sys.argv[1]
    ruta_xlsx = sys.argv[2]
    main(ruta_json, ruta_xlsx)
