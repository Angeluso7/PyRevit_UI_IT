# -*- coding: utf-8 -*-
"""
leer_xlsm_codigos.py — CPython 3
Lee hoja 'CODIGO' del .xlsm, genera CODIGO.csv en data_output_dir y
devuelve su ruta por stdout (última línea).

Uso:
    python leer_xlsm_codigos.py <ruta_xlsm> <data_output_dir>

Salida stdout (última línea):
    <ruta_absoluta_a_CODIGO.csv>
"""

import sys
import os
import csv
import openpyxl


def leer_hoja_codigo(ruta_xlsm, data_dir):
    """
    Lee hoja 'CODIGO' desde fila 1 (encabezados) + fila 2 en adelante (datos).
    Si la hoja no tiene encabezados en fila 1 se generan nombres genéricos.
    Genera CODIGO.csv en data_dir con delimitador ',' (compatible con csv.DictReader)
    y devuelve la ruta del CSV.
    """
    if not os.path.isfile(ruta_xlsm):
        raise RuntimeError("Archivo no encontrado: {}".format(ruta_xlsm))

    wb = openpyxl.load_workbook(ruta_xlsm, data_only=True, read_only=True)

    # Buscar hoja CODIGO (case-insensitive)
    hoja = None
    for nombre in wb.sheetnames:
        if nombre.strip().upper() == "CODIGO":
            hoja = wb[nombre]
            break
    if hoja is None:
        raise RuntimeError(
            "No se encontró hoja 'CODIGO'. Hojas disponibles: {}".format(
                ", ".join(wb.sheetnames)
            )
        )

    # Leer todas las filas como listas (incluyendo fila 1)
    todas_las_filas = []
    for row in hoja.iter_rows(values_only=True):
        todas_las_filas.append(list(row))

    wb.close()

    if not todas_las_filas:
        raise RuntimeError("Hoja CODIGO está vacía.")

    # ── Determinar encabezados ────────────────────────────────
    # Si la primera celda de la fila 1 parece texto (no un código como CM01-xxx),
    # se usa como fila de encabezados. Caso contrario se generan nombres genéricos.
    primera_celda = str(todas_las_filas[0][0] or "").strip()
    es_encabezado = primera_celda and not (
        len(primera_celda) >= 4
        and primera_celda[:2].upper() == "CM"
        and primera_celda[2:4].isdigit()
    )

    if es_encabezado:
        fila_headers = todas_las_filas[0]
        filas_datos  = todas_las_filas[1:]
    else:
        fila_headers = None
        filas_datos  = todas_las_filas

    # Filtrar filas completamente vacías y las que no tengan CodIntBIM
    filas_utiles = [
        r for r in filas_datos
        if r and r[0] not in (None, "")
    ]

    if not filas_utiles:
        raise RuntimeError("Hoja CODIGO sin filas de datos útiles.")

    max_cols = max(len(r) for r in filas_utiles)

    # Construir headers definitivos
    if fila_headers:
        headers = []
        for i, h in enumerate(fila_headers[:max_cols]):
            val = str(h or "").strip()
            headers.append(val if val else "Col{}".format(i + 1))
        # Si hay más columnas en datos que en encabezados, extender
        while len(headers) < max_cols:
            headers.append("Col{}".format(len(headers) + 1))
    else:
        # Sin encabezados: primera columna = CodIntBIM, resto genéricos
        headers = ["CodIntBIM"] + ["Col{}".format(i) for i in range(2, max_cols + 1)]

    # Asegurar que primera columna se llame CodIntBIM si no tiene nombre claro
    if headers[0].upper() in ("", "CODINTBIM", "COD_INT_BIM", "COD INT BIM", "CODIGO"):
        headers[0] = "CodIntBIM"

    # ── Escribir CSV con delimitador ',' ──────────────────────
    if not os.path.exists(data_dir):
        os.makedirs(data_dir)

    csv_path = os.path.join(data_dir, "CODIGO.csv")
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        # utf-8-sig para compatibilidad con Excel al abrir manualmente
        writer = csv.writer(f, delimiter=",", quoting=csv.QUOTE_MINIMAL)
        writer.writerow(headers)
        for row in filas_utiles:
            # Rellenar con vacíos si la fila tiene menos columnas que max_cols
            row_padded = list(row) + [""] * (max_cols - len(row))
            # Convertir None a cadena vacía
            row_limpia = [
                "" if v is None else str(v).strip()
                for v in row_padded[:max_cols]
            ]
            writer.writerow(row_limpia)

    print(csv_path)
    return csv_path


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print(
            "Uso: leer_xlsm_codigos.py <ruta_xlsm> <data_output_dir>",
            file=sys.stderr
        )
        sys.exit(1)

    ruta_xlsm = sys.argv[1].strip('"').strip("'")
    data_dir  = sys.argv[2].strip('"').strip("'")

    try:
        leer_hoja_codigo(ruta_xlsm, data_dir)
    except Exception as e:
        print("Error: {}".format(e), file=sys.stderr)
        sys.exit(1)