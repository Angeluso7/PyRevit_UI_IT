# -*- coding: utf-8 -*-
"""
scripts_cpython/ui_comparacion.py  v1.7
Cambios v1.5:
- [FIX] leer_xlsm_codigos: keep_vba=False + manejo robusto de xlsm con macros.
- [FIX] exportar_json_y_formatear: clave "elementos_por_tabla" (no "datos_por_tabla").
- [FIX] headers_por_tabla incluido en comparacion.json con clave correcta.

Cambios v1.6:
- [FIX] Combobox muestra nombre de planilla legible en lugar del codigo CM raw.
- [ADD] _etiqueta_cm() y _mapa_etiqueta_clave.

Cambios v1.7:
- [FIX] mostrar_ui y __main__ completos (el archivo estaba truncado en repo).
- [FIX] CREATE_NO_WINDOW en subprocess para evitar consola flash en Windows.
"""

import sys
import os
import json
import subprocess
import traceback
import tkinter as tk
from tkinter import ttk, messagebox

COLOR_ESTADOS = {
    'ok':           '#C6EFCE',
    'falta_modelo': '#FFC7CE',
    'falta_excel':  '#FFEB9C',
    'difiere':      '#F4B084',
    'no_existe':    '#66FFFF',
}

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_EXT_ROOT   = os.path.abspath(os.path.join(_SCRIPT_DIR, '..'))

DATA_DIR_EXT           = os.path.join(_EXT_ROOT, 'data')
PLANILLAS_HEADERS_JSON = os.path.join(DATA_DIR_EXT, 'master', 'planillas_headers_order.json')
CONFIG_PROYECTO_ACTIVO = os.path.join(DATA_DIR_EXT, 'master', 'config_proyecto_activo.json')
PROYECTOS_DIR          = os.path.join(DATA_DIR_EXT, 'proyectos')
UI_LOG_PATH            = os.path.join(DATA_DIR_EXT, 'logs', 'ui_comparacion_log.txt')
HEADER_VINCULO         = u"V\u00ednculo RVT: Nombre de archivo"

CREATE_NO_WINDOW = 0x08000000 if sys.platform == 'win32' else 0


def log(msg):
    try:
        from datetime import datetime
        ts    = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        linea = u"[{}] {}\n".format(ts, msg)
        os.makedirs(os.path.dirname(UI_LOG_PATH), exist_ok=True)
        with open(UI_LOG_PATH, "a", encoding="utf-8") as f:
            f.write(linea)
    except Exception:
        pass


def log_exc(contexto):
    try:
        tb = traceback.format_exc()
        log("{} -> {}".format(contexto, tb))
    except Exception:
        pass


# -- JSON ----------------------------------------------------------------------
def cargar_json(ruta, default):
    try:
        if not os.path.exists(ruta):
            log("cargar_json: no existe -> {}".format(ruta))
            return default
        with open(ruta, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        log_exc("cargar_json {}".format(ruta))
        return default


# -- Repo activo ---------------------------------------------------------------
def get_repo_activo_path():
    cfg  = cargar_json(CONFIG_PROYECTO_ACTIVO, {})
    ruta = (cfg.get("ruta_repositorio_activo") or "").strip()
    if ruta and os.path.exists(ruta):
        return ruta
    nup = (cfg.get("nup_activo") or "").strip()
    if nup:
        return os.path.join(PROYECTOS_DIR, u'repositorio_datos_{}.json'.format(nup))
    return ""


def cargar_repo_activo():
    ruta = get_repo_activo_path()
    if not ruta or not os.path.exists(ruta):
        log("cargar_repo_activo: no disponible -> {}".format(ruta))
        return {}
    repo = cargar_json(ruta, {})
    log("cargar_repo_activo: {} registros".format(len(repo)))
    return repo


# -- Leer xlsm con openpyxl ---------------------------------------------------
def leer_xlsm_codigos(ruta_xlsm):
    """
    Lee todas las hojas del .xlsm con openpyxl.
    keep_vba=False -> ignora macros VBA.
    data_only=True -> lee valores calculados.
    read_only=True -> apertura rapida.
    """
    try:
        import openpyxl
    except ImportError:
        log("leer_xlsm_codigos: openpyxl no instalado")
        messagebox.showerror(
            "Dependencia faltante",
            "openpyxl no esta instalado.\nEjecuta: pip install openpyxl"
        )
        raise

    if not os.path.exists(ruta_xlsm):
        raise FileNotFoundError("No se encontro el archivo: {}".format(ruta_xlsm))

    log("leer_xlsm_codigos: ruta={}".format(ruta_xlsm))
    filas = []
    try:
        wb = openpyxl.load_workbook(ruta_xlsm, read_only=True, data_only=True, keep_vba=False)
        for nombre_hoja in wb.sheetnames:
            try:
                ws = wb[nombre_hoja]
                for row in ws.iter_rows(values_only=True):
                    if not row:
                        continue
                    primera = str(row[0] or '').strip()
                    if len(primera) >= 4 and primera[:2] == 'CM':
                        filas.append([str(c or '').strip() for c in row])
            except Exception as e_hoja:
                log("leer_xlsm_codigos: error hoja '{}' -> {}".format(nombre_hoja, e_hoja))
                continue
        wb.close()
        log("leer_xlsm_codigos: {} filas CM".format(len(filas)))
        return filas
    except Exception:
        log_exc("leer_xlsm_codigos")
        raise


# -- Headers ------------------------------------------------------------------
def _normalizar_headers(headers_raw):
    if not headers_raw:
        return []
    headers = [h for h in headers_raw if h != HEADER_VINCULO]
    if "CodIntBIM" in headers:
        headers = ["CodIntBIM"] + [h for h in headers if h != "CodIntBIM"]
    else:
        headers = ["CodIntBIM"] + headers
    return headers


def cargar_headers_planilla(nombre_planilla, codigo_cm):
    if not os.path.exists(PLANILLAS_HEADERS_JSON):
        return None
    try:
        with open(PLANILLAS_HEADERS_JSON, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception:
        log_exc("cargar_headers_planilla {}".format(PLANILLAS_HEADERS_JSON))
        return None
    clave   = "{}::{}".format(nombre_planilla, codigo_cm)
    headers = data.get(clave) or data.get(codigo_cm)
    if headers:
        h_norm = _normalizar_headers(headers)
        log("cargar_headers_planilla: {} -> {} cols".format(nombre_planilla, len(h_norm)))
        return h_norm
    return None


# -- Construir Excel por planilla ---------------------------------------------
def construir_excel_por_planilla(filas_xlsm, codigos_planillas):
    planilla_a_cm = {}
    for k, v in codigos_planillas.items():
        try:
            if isinstance(k, str) and len(k) == 4 and k.startswith("CM"):
                planilla_a_cm[k] = v
            elif isinstance(v, str) and len(v) == 4 and v.startswith("CM"):
                planilla_a_cm[v] = k
        except Exception:
            pass

    log("construir_excel_por_planilla: planilla_a_cm -> {}".format(planilla_a_cm))
    datos_excel_planilla = {}

    for row in filas_xlsm:
        if not row:
            continue
        primera_celda = (row[0] or '').strip()
        if not primera_celda or len(primera_celda) < 4:
            continue
        pref = primera_celda[:4]
        nombre_planilla = planilla_a_cm.get(pref)
        if not nombre_planilla:
            continue
        headers = cargar_headers_planilla(nombre_planilla, pref)
        if not headers:
            headers = _normalizar_headers(["Pa{}".format(i) for i in range(1, len(row) + 1)])
        if nombre_planilla not in datos_excel_planilla:
            datos_excel_planilla[nombre_planilla] = {"codigo_cm": pref, "headers": headers, "filas": []}
        bloque  = datos_excel_planilla[nombre_planilla]
        valores = [row[idx_h] if idx_h < len(row) else '' for idx_h in range(len(headers))]
        codint  = (valores[0] or '').strip()
        bloque["filas"].append({"CodIntBIM": codint, "valores": valores})

    log("construir_excel_por_planilla: {} planillas".format(len(datos_excel_planilla)))
    return datos_excel_planilla


# -- Modelo -------------------------------------------------------------------
def leer_modelo_por_cm(path_json_modelo):
    try:
        if not os.path.exists(path_json_modelo):
            log("leer_modelo_por_cm: no existe -> {}".format(path_json_modelo))
            return {}
        with open(path_json_modelo, 'r', encoding='utf-8') as f:
            data = json.load(f)
        result = {cm: lista for cm, lista in data.items() if isinstance(lista, list)}
        log("leer_modelo_por_cm: CMs -> {}".format(list(result.keys())))
        return result
    except Exception:
        log_exc("leer_modelo_por_cm")
        raise


def enriquecer_modelo_con_bd(datos_modelo_cm, repo):
    if not repo:
        return datos_modelo_cm
    indice_bd = {}
    for k, v in repo.items():
        eid = str(v.get("ElementId", "")).strip()
        cod = (v.get("CodIntBIM") or "").strip()
        if eid and cod:
            indice_bd[(eid, cod)] = v
    nuevo = {}
    for cm, lista in datos_modelo_cm.items():
        nueva_lista = []
        for fila in lista:
            cod    = (fila.get("CodIntBIM") or '').strip()
            eid    = str(fila.get("ElementId", "")).strip()
            bd_row = indice_bd.get((eid, cod))
            if bd_row:
                combinado = dict(fila)
                combinado.update(bd_row)
                nueva_lista.append(combinado)
            else:
                nueva_lista.append(fila)
        nuevo[cm] = nueva_lista
    return nuevo


# -- Comparacion --------------------------------------------------------------
def _norm_val(v):
    v = (v or '').strip()
    return '' if v == '-' else v


def _get_val(row_dict, header_name):
    if not row_dict or header_name == HEADER_VINCULO:
        return ''
    v = (row_dict.get(header_name) or '').strip()
    return '' if v == '-' else v


def _comparar_por_planilla(codigo_cm, headers, filas_excel, filas_modelo):
    mapa_excel  = {}
    for fila in filas_excel:
        c = (fila["CodIntBIM"] or '').strip()
        if c:
            mapa_excel.setdefault(c, []).append(fila)

    mapa_modelo = {}
    for row_m in filas_modelo:
        c = (row_m.get("CodIntBIM") or '').strip()
        if c:
            mapa_modelo.setdefault(c, []).append(row_m)

    filas_tabla_excel  = []
    filas_tabla_modelo = []
    filas_fusionadas   = []
    codigos_union      = sorted(set(mapa_excel.keys()) | set(mapa_modelo.keys()))

    for cod in codigos_union:
        lista_excel  = mapa_excel.get(cod, [])
        lista_modelo = mapa_modelo.get(cod, [])
        fila_x0 = lista_excel[0]  if lista_excel  else None
        fila_m0 = lista_modelo[0] if lista_modelo else None

        valores_fusion = []
        estados_fusion = []
        for h in headers:
            vx_raw = ''
            vm_raw = ''
            if fila_x0:
                idx_h  = headers.index(h)
                vals   = fila_x0["valores"]
                vx_raw = vals[idx_h] if idx_h < len(vals) else ''
            if fila_m0:
                vm_raw = _get_val(fila_m0, h)
            vx = _norm_val(vx_raw)
            vm = _norm_val(vm_raw)

            if not vx and not vm:
                valores_fusion.append('');      estados_fusion.append('no_existe')
            elif vx == vm:
                valores_fusion.append(vx_raw);  estados_fusion.append('ok')
            elif vx and not vm:
                valores_fusion.append(vx_raw);  estados_fusion.append('falta_modelo')
            elif not vx and vm:
                valores_fusion.append(vm_raw);  estados_fusion.append('falta_excel')
            else:
                valores_fusion.append(u"Planilla: {}\nModelo: {}".format(vx_raw, vm_raw))
                estados_fusion.append('difiere')

        filas_fusionadas.append({
            "CodIntBIM": cod, "valores": valores_fusion,
            "estado_por_celda": estados_fusion
        })

        for fila_x in lista_excel:
            vals_x = fila_x["valores"]
            valores, estados = [], []
            fila_m = lista_modelo[0] if lista_modelo else None
            for idx, h in enumerate(headers):
                vx_raw = vals_x[idx] if idx < len(vals_x) else ''
                vx     = _norm_val(vx_raw)
                vm_raw = _get_val(fila_m, h) if fila_m else ''
                vm     = _norm_val(vm_raw)
                if not vx and not vm:
                    valores.append('');   estados.append('no_existe')
                elif vx == vm:
                    valores.append(vx_raw); estados.append('ok')
                elif vx and not vm:
                    valores.append(vx_raw); estados.append('falta_modelo')
                elif not vx and vm:
                    valores.append(u"Planilla: {}\nModelo: {}".format(vx_raw, vm_raw))
                    estados.append('falta_excel')
                else:
                    valores.append(u"Planilla: {}\nModelo: {}".format(vx_raw, vm_raw))
                    estados.append('difiere')
            filas_tabla_excel.append({
                "origen": "excel", "CodIntBIM": cod,
                "valores": valores, "estado_por_celda": estados
            })

        for fila_m in lista_modelo:
            valores, estados = [], []
            fila_x = lista_excel[0] if lista_excel else None
            for idx, h in enumerate(headers):
                vm_raw = _get_val(fila_m, h)
                vm     = _norm_val(vm_raw)
                vals_x = fila_x["valores"] if fila_x else []
                vx_raw = vals_x[idx] if idx < len(vals_x) else ''
                vx     = _norm_val(vx_raw)
                if not vx and not vm:
                    valores.append('');   estados.append('no_existe')
                elif vx == vm:
                    valores.append(vm_raw); estados.append('ok')
                elif vm and not vx:
                    valores.append(vm_raw); estados.append('falta_excel')
                elif not vm and vx:
                    valores.append(u"Modelo: {}\nPlanilla: {}".format(vm_raw, vx_raw))
                    estados.append('falta_modelo')
                else:
                    valores.append(u"Modelo: {}\nPlanilla: {}".format(vm_raw, vx_raw))
                    estados.append('difiere')
            filas_tabla_modelo.append({
                "origen": "modelo", "CodIntBIM": cod,
                "valores": valores, "estado_por_celda": estados
            })

    return filas_tabla_excel, filas_tabla_modelo, filas_fusionadas


def construir_tabla_comparativa(datos_excel_planilla, datos_modelo_cm, codigos_planillas):
    datos_comparacion = {}
    cm_to_planilla    = {}
    for k, v in codigos_planillas.items():
        try:
            if isinstance(v, str) and len(v) == 4 and v.startswith("CM"):
                cm_to_planilla[v] = k
            elif isinstance(k, str) and len(k) == 4 and k.startswith("CM"):
                cm_to_planilla[k] = v
        except Exception:
            pass

    for planilla, bloque_excel in datos_excel_planilla.items():
        codigo_cm    = bloque_excel['codigo_cm']
        headers      = bloque_excel['headers']
        filas_excel  = bloque_excel['filas']
        filas_modelo = datos_modelo_cm.get(codigo_cm, [])
        fe, fm, ff   = _comparar_por_planilla(codigo_cm, headers, filas_excel, filas_modelo)
        datos_comparacion[planilla] = {
            "codigo_cm": codigo_cm, "headers": headers,
            "filas_excel_base": fe, "filas_modelo_base": fm, "filas_fusionadas": ff
        }

    for codigo_cm, lista_modelo in datos_modelo_cm.items():
        if any(i["codigo_cm"] == codigo_cm for i in datos_comparacion.values()):
            continue
        nombre_planilla = cm_to_planilla.get(codigo_cm, codigo_cm)
        headers = cargar_headers_planilla(nombre_planilla, codigo_cm)
        if not headers:
            keys    = sorted(set(k for fila in lista_modelo for k in fila.keys())) if lista_modelo else []
            keys    = [k for k in keys if k != HEADER_VINCULO]
            headers = _normalizar_headers(["CodIntBIM"] + [k for k in keys if k != "CodIntBIM"])
        fe, fm, ff = _comparar_por_planilla(codigo_cm, headers, [], lista_modelo)
        datos_comparacion[nombre_planilla] = {
            "codigo_cm": codigo_cm, "headers": headers,
            "filas_excel_base": fe, "filas_modelo_base": fm, "filas_fusionadas": ff
        }

    log("construir_tabla_comparativa: {} planillas".format(len(datos_comparacion)))
    return datos_comparacion


# -- Exportar Excel -----------------------------------------------------------
def exportar_json_y_formatear(datos_comparacion, data_dir,
                              formatear_script, ruta_xlsx_salida, python_exe):
    try:
        elementos_por_tabla = {}
        headers_por_tabla   = {}
        codigos, nombres    = [], []

        for nombre_planilla, bloque in datos_comparacion.items():
            codigo_cm = bloque['codigo_cm']
            headers   = bloque['headers']
            filas     = bloque.get('filas_fusionadas', [])
            elementos = []
            for fila in filas:
                vals = fila.get('valores', [])
                elem = {}
                for idx, h in enumerate(headers):
                    elem[h] = vals[idx] if idx < len(vals) else ''
                elementos.append(elem)
            elementos_por_tabla[codigo_cm] = elementos
            headers_por_tabla[codigo_cm]   = headers
            codigos.append(codigo_cm)
            nombres.append(nombre_planilla)

        pares             = sorted(zip(codigos, nombres), key=lambda x: x[0])
        codigos_ordenados = [p[0] for p in pares]
        nombres_ordenados = [p[1] for p in pares]

        datos_export = {
            "listado_tablas": {"valores": codigos_ordenados, "claves": nombres_ordenados},
            "elementos_por_tabla": elementos_por_tabla,
            "headers_por_tabla":   headers_por_tabla,
        }

        os.makedirs(data_dir, exist_ok=True)
        ruta_json = os.path.join(data_dir, "comparacion.json")
        with open(ruta_json, "w", encoding="utf-8") as f:
            json.dump(datos_export, f, ensure_ascii=False, indent=2)
        log("exportar_json_y_formatear: JSON -> {}".format(ruta_json))

        cmd    = [python_exe, formatear_script, ruta_json, ruta_xlsx_salida]
        result = subprocess.run(
            cmd, capture_output=True, text=True, timeout=120,
            creationflags=CREATE_NO_WINDOW
        )
        if result.returncode != 0:
            log("exportar_json_y_formatear: ERROR -> {}".format(result.stderr))
            messagebox.showerror(
                "Error al exportar",
                "El formateador termino con error:\n{}".format(result.stderr[:500])
            )
        else:
            log("exportar_json_y_formatear: OK -> {}".format(ruta_xlsx_salida))
            messagebox.showinfo("Exportado", u"Archivo generado:\n{}".format(ruta_xlsx_salida))

    except Exception:
        log_exc("exportar_json_y_formatear")
        messagebox.showerror(
            "Error al exportar",
            "Error inesperado al generar el Excel.\nRevisa ui_comparacion_log.txt."
        )


# -- Treeview helpers ---------------------------------------------------------
def limpiar_treeview(tree):
    for item in tree.get_children():
        tree.delete(item)
    tree["columns"] = []


def crear_treeview_tabla(parent):
    frame = ttk.Frame(parent)
    frame.pack(fill=tk.BOTH, expand=True)
    vsb  = ttk.Scrollbar(frame, orient=tk.VERTICAL)
    hsb  = ttk.Scrollbar(frame, orient=tk.HORIZONTAL)
    tree = ttk.Treeview(frame, show="headings",
                        yscrollcommand=vsb.set, xscrollcommand=hsb.set)
    vsb.config(command=tree.yview)
    hsb.config(command=tree.xview)
    vsb.pack(side=tk.RIGHT,  fill=tk.Y)
    hsb.pack(side=tk.BOTTOM, fill=tk.X)
    tree.pack(fill=tk.BOTH, expand=True)
    for estado, color in COLOR_ESTADOS.items():
        tree.tag_configure(estado, background=color)
    return tree


def poblar_treeview(tree, headers, filas_tabla):
    limpiar_treeview(tree)
    if not headers:
        return
    columnas = list(headers)
    tree["columns"] = columnas
    for h in columnas:
        tree.heading(h, text=h)
        tree.column(h, width=200, stretch=False, anchor="w")
    if not filas_tabla:
        return

    def peor_estado(estados):
        for e in ('difiere', 'falta_modelo', 'falta_excel', 'no_existe'):
            if e in estados:
                return e
        return 'ok'

    for fila in filas_tabla:
        valores = list(fila.get('valores', []))
        estados = list(fila.get('estado_por_celda') or [])
        if len(valores) < len(columnas):
            valores += [""] * (len(columnas) - len(valores))
        if len(estados) < len(columnas):
            estados += ['ok'] * (len(columnas) - len(estados))
        tree.insert('', 'end', values=valores[:len(columnas)],
                    tags=(peor_estado(estados),))


# -- UI principal -------------------------------------------------------------
def mostrar_ui(datos_comparacion, on_exportar):
    root = tk.Tk()
    root.title(u"Planilla vs Modelo")
    root.geometry("1200x650")
    root.minsize(800, 450)

    export_realizado = {"value": False}
    style = ttk.Style(root)
    style.configure("Treeview", rowheight=30)
    style.configure("Treeview.Heading", font=("TkDefaultFont", 9, "bold"))

    # Frame superior: selector de planilla
    frame_main = ttk.Frame(root)
    frame_main.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

    frame_top = ttk.Frame(frame_main)
    frame_top.pack(fill=tk.X, padx=5, pady=(5, 2))
    ttk.Label(frame_top, text=u"Tabla:").pack(side=tk.LEFT, padx=(0, 5))

    def _etiqueta_cm(clave):
        if "::" in clave:
            partes = clave.split("::", 1)
            return u"{}  ({})".format(partes[0].strip(), partes[1].strip())
        return clave

    claves_ordenadas     = sorted(datos_comparacion.keys())
    etiquetas            = [_etiqueta_cm(c) for c in claves_ordenadas]
    _mapa_etiqueta_clave = dict(zip(etiquetas, claves_ordenadas))

    selected_nombre = tk.StringVar()
    combo_tablas = ttk.Combobox(
        frame_top, textvariable=selected_nombre,
        values=etiquetas, state="readonly", width=55
    )
    combo_tablas.pack(side=tk.LEFT, fill=tk.X, expand=True)

    lbl_conteo = ttk.Label(frame_top, text="", foreground="gray")
    lbl_conteo.pack(side=tk.RIGHT, padx=8)

    # Notebook con dos tabs
    ttk.Label(frame_main, text=u"Tabla comparativa",
              font=("TkDefaultFont", 10, "bold")).pack(anchor=tk.W, padx=5, pady=(2, 0))

    frame_center = ttk.Frame(frame_main)
    frame_center.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

    notebook     = ttk.Notebook(frame_center)
    notebook.pack(fill=tk.BOTH, expand=True)
    frame_excel  = ttk.Frame(notebook)
    frame_modelo = ttk.Frame(notebook)
    notebook.add(frame_excel,  text=u"Desde Excel")
    notebook.add(frame_modelo, text=u"Desde Modelo")

    tree_excel  = crear_treeview_tabla(frame_excel)
    tree_modelo = crear_treeview_tabla(frame_modelo)

    # Frame inferior: leyenda + botones
    frame_bottom = ttk.Frame(root)
    frame_bottom.pack(fill=tk.X, padx=5, pady=5, side=tk.BOTTOM)

    frame_leyenda = ttk.Frame(frame_bottom)
    frame_leyenda.pack(side=tk.LEFT, padx=5)
    ttk.Label(frame_leyenda, text=u"Leyenda:").pack(side=tk.LEFT, padx=(0, 4))
    LABELS_ESTADO = {
        'ok':           u'OK / Coincide',
        'falta_modelo': u'Falta en Modelo',
        'falta_excel':  u'Falta en Excel',
        'no_existe':    u'No existe',
        'difiere':      u'Diferencia de valor',
    }
    for estado, color in COLOR_ESTADOS.items():
        tk.Label(frame_leyenda, width=2, background=color,
                 relief=tk.SUNKEN, bd=1).pack(side=tk.LEFT, padx=2)
        ttk.Label(frame_leyenda, text=LABELS_ESTADO.get(estado, estado),
                  font=("TkDefaultFont", 8)).pack(side=tk.LEFT, padx=(0, 4))

    frame_botones = ttk.Frame(frame_bottom)
    frame_botones.pack(side=tk.RIGHT)

    # Funciones internas
    def actualizar_tablas(etiqueta):
        clave  = _mapa_etiqueta_clave.get(etiqueta, etiqueta)
        bloque = datos_comparacion.get(clave)
        if not bloque:
            limpiar_treeview(tree_excel)
            limpiar_treeview(tree_modelo)
            lbl_conteo.config(text="")
            return
        headers      = bloque.get('headers', [])
        filas_excel  = bloque.get('filas_excel_base',  [])
        filas_modelo = bloque.get('filas_modelo_base', [])
        poblar_treeview(tree_excel,  headers, filas_excel)
        poblar_treeview(tree_modelo, headers, filas_modelo)
        lbl_conteo.config(
            text=u"Excel: {} filas  |  Modelo: {} filas".format(
                len(filas_excel), len(filas_modelo))
        )

    combo_tablas.bind("<<ComboboxSelected>>",
                      lambda e: actualizar_tablas(selected_nombre.get()))

    def cmd_exportar():
        on_exportar()
        export_realizado["value"] = True

    def cmd_salir():
        if not export_realizado["value"]:
            if not messagebox.askyesno(
                u"Salir",
                u"No has exportado el archivo.\n\u00bfSeguro que quieres salir sin exportar?"
            ):
                return
        root.destroy()

    ttk.Button(frame_botones, text=u"Exportar a Excel",
               command=cmd_exportar).pack(side=tk.LEFT, padx=5)
    ttk.Button(frame_botones, text=u"Salir",
               command=cmd_salir).pack(side=tk.LEFT, padx=5)
    ttk.Sizegrip(root).pack(side=tk.BOTTOM, anchor=tk.SE)

    # Seleccion inicial
    if etiquetas:
        selected_nombre.set(etiquetas[0])
        actualizar_tablas(etiquetas[0])
    else:
        messagebox.showwarning(
            u"Sin datos",
            u"No se encontraron planillas para comparar.\n"
            u"Verifica que el archivo Excel contiene codigos CM y que "
            u"script.json tiene codigos_planillas configurados."
        )

    root.protocol("WM_DELETE_WINDOW", cmd_salir)
    root.mainloop()


# -- MAIN ---------------------------------------------------------------------
if __name__ == '__main__':
    try:
        log("==== Inicio ui_comparacion.py v1.7 ====")

        if len(sys.argv) < 9:
            log("main: argumentos insuficientes -> {}".format(sys.argv))
            print(
                "Uso: ui_comparacion.py "
                "<script_json> <ruta_xlsm> <data_dir> "
                "<formatear_script> <ruta_xlsx_salida> "
                "<python_exe> <modelo_json> <headers_json>",
                file=sys.stderr
            )
            sys.exit(120)

        script_json      = sys.argv[1]
        ruta_xlsm        = sys.argv[2]
        data_dir         = sys.argv[3]
        formatear_script = sys.argv[4]
        ruta_xlsx_salida = sys.argv[5]
        python_exe       = sys.argv[6]
        modelo_json      = sys.argv[7]
        headers_json     = sys.argv[8]

        log("Args: script_json={} ruta_xlsm={} data_dir={} "
            "formatear={} salida={} py={} modelo={}".format(
            script_json, ruta_xlsm, data_dir,
            formatear_script, ruta_xlsx_salida, python_exe, modelo_json))

        cfg               = cargar_json(script_json, {})
        codigos_planillas = cfg.get('codigos_planillas', {})
        if not codigos_planillas:
            log("main: script.json no tiene codigos_planillas")
            messagebox.showerror(
                u"Configuracion incompleta",
                u"script.json no tiene la clave 'codigos_planillas'.\n"
                u"Ruta: {}".format(script_json)
            )
            sys.exit(1)

        filas_xlsm           = leer_xlsm_codigos(ruta_xlsm)
        datos_excel_planilla = construir_excel_por_planilla(filas_xlsm, codigos_planillas)

        datos_modelo_cm_raw  = leer_modelo_por_cm(modelo_json)
        repo_activo          = cargar_repo_activo()
        datos_modelo_cm      = enriquecer_modelo_con_bd(datos_modelo_cm_raw, repo_activo)

        datos_comparacion = construir_tabla_comparativa(
            datos_excel_planilla, datos_modelo_cm, codigos_planillas
        )
        log("main: {} planillas en datos_comparacion".format(len(datos_comparacion)))

        if not datos_comparacion:
            messagebox.showwarning(
                u"Sin datos",
                u"No se encontraron filas en el dataset.\n\n"
                u"Posibles causas:\n"
                u"  1. El archivo Excel no tiene celdas que empiecen con 'CM'.\n"
                u"  2. Los codigos CM del Excel no coinciden con codigos_planillas en script.json.\n"
                u"  3. El modelo Revit no tenia elementos con CodIntBIM.\n\n"
                u"Revisa ui_comparacion_log.txt para mas detalles."
            )
            sys.exit(0)

        def on_export():
            exportar_json_y_formatear(
                datos_comparacion, data_dir,
                formatear_script, ruta_xlsx_salida, python_exe
            )

        mostrar_ui(datos_comparacion, on_export)
        log("==== Fin ui_comparacion.py v1.7 ====")

    except SystemExit as se:
        log("SystemExit -> {}".format(se))
        raise
    except Exception:
        log_exc("ui_comparacion main error general")
        try:
            messagebox.showerror(
                u"Error",
                u"Error inesperado en ui_comparacion.py.\nRevisa ui_comparacion_log.txt."
            )
        except Exception:
            pass
        sys.exit(120)
