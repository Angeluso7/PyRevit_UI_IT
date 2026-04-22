# -*- coding: utf-8 -*-
"""
formatear_revision_xlsx.py - CPython 3
Genera el Excel de Revision de Elementos con formato vertical por parametro.

Uso:
    python formatear_revision_xlsx.py <temp_datos.json> <ruta_xlsx_salida> <data_master_dir>

Argumentos:
    sys.argv[1]  = _temp_datos.json  (generado por extraer_modelos_bim.py)
    sys.argv[2]  = ruta .xlsx de salida
    sys.argv[3]  = carpeta data/master (para hallar colores_parametros.json)

Formato VERTICAL con colores por parametro:
  - Columna A: nombres de parametros (filas)
  - Columnas B, C, D...: valores por elemento
  - Colores desde colores_parametros.json

Reglas de color:
  1. CodIntBIM           -> D7D7D7 (gris medio)
  2. Parametros extra    -> D7D7D7 (gris medio)
  3. En colores_json     -> color del JSON
  4. NO en colores_json  -> FFFFCC (amarillo claro)
"""

import sys
import os
import json
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Colores base ─────────────────────────────────────────────
COLOR_DEFAULT_NO_JSON = "FFFFFFCC"   # amarillo claro: parametro sin color asignado
COLOR_ADICIONALES     = "FFD7D7D7"   # gris medio: campos extra (ElementId, Categoria, etc.)
COLOR_CODINTBIM       = "FFD7D7D7"   # gris medio: CodIntBIM
COLOR_HEADER_INDICE   = "FF366092"   # azul oscuro: encabezados indice/excepciones
FONT_COLOR_INDICE     = "FFFFFFFF"   # blanco
BORDER_STYLE = Side(style="thin", color="FF000000")

# ── Helpers generales ─────────────────────────────────────────
def cargar_json(ruta, default=None):
    if default is None:
        default = {}
    try:
        with open(ruta, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        print("AVISO cargar_json {}: {}".format(ruta, e))
        return default

def cargar_colores_parametros(data_master_dir):
    ruta = os.path.join(data_master_dir, "colores_parametros.json")
    if not os.path.exists(ruta):
        print("AVISO: no se encontro colores_parametros.json en: {}".format(ruta))
        return {}
    try:
        with open(ruta, "r", encoding="utf-8") as f:
            data = json.load(f)
        return {k.strip(): v.strip() for k, v in data.items()}
    except Exception as e:
        print("Error cargando colores: {}".format(e))
        return {}

# ── Estilos ───────────────────────────────────────────────────
def aplicar_estilo_celda(cell, color_hex, bold=False, fuente_color="FF000000"):
    cell.fill      = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
    cell.font      = Font(bold=bold, color=fuente_color)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border    = Border(
        left=BORDER_STYLE, right=BORDER_STYLE,
        top=BORDER_STYLE,  bottom=BORDER_STYLE
    )

def autoajustar_columnas(ws, max_ancho=60):
    for col_idx in range(1, ws.max_column + 1):
        max_len  = 0
        col_letter = get_column_letter(col_idx)
        for row_idx in range(1, ws.max_row + 1):
            v = ws.cell(row_idx, col_idx).value
            if v:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 15), max_ancho)

def autoajustar_filas(ws, alto=20):
    for row_idx in range(1, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = alto

# ── Color de parametro ────────────────────────────────────────
PARAMETROS_ADICIONALES = [
    "ElementId",
    u"Categor\u00eda",
    "Categoria",
    "Familia",
    "Tipo",
    "Nombre_RVT",
    u"Situaci\u00f3n CodIntBIM",
    "Situacion CodIntBIM",
    "Elementos",
]

def obtener_color_parametro(param, colores_dict):
    if param == "CodIntBIM":
        return COLOR_CODINTBIM
    if param in PARAMETROS_ADICIONALES:
        return COLOR_ADICIONALES
    if param in colores_dict:
        return colores_dict[param]
    return COLOR_DEFAULT_NO_JSON

# ── Hoja Indice ───────────────────────────────────────────────
def generar_hoja_indice(ws, listado_tablas):
    ws.title = "Indice"

    ws["B3"] = "Listado de Tablas"
    ws["B3"].font = Font(bold=True, size=12)

    for col, h in [("B", "Codigo"), ("C", "Nombre Tabla (Schedule)")]:
        c = ws[col + "4"]
        c.value = h
        aplicar_estilo_celda(c, COLOR_HEADER_INDICE, bold=True, fuente_color=FONT_COLOR_INDICE)

    valores = listado_tablas.get("valores", [])
    claves  = listado_tablas.get("claves",  [])
    pares   = sorted(zip(valores, claves), key=lambda x: x[0])

    for idx, (codigo, nombre_schedule) in enumerate(pares):
        fila = 5 + idx
        ws["B{}".format(fila)] = codigo
        ws["C{}".format(fila)] = nombre_schedule

    autoajustar_columnas(ws)

# ── Hoja Excepciones ──────────────────────────────────────────
def generar_hoja_excepciones(ws, excepciones):
    ws.title = "Excepciones"

    ws["A1"] = "Excepciones durante la extraccion"
    ws["A1"].font = Font(bold=True, size=12)

    campos = ["ElementId", "CodIntBIM", u"Categor\u00eda",
              "Familia", "Tipo", "Nombre_RVT", u"Situaci\u00f3n"]

    for col_idx, campo in enumerate(campos, 1):
        c = ws.cell(1, col_idx, value=campo)
        aplicar_estilo_celda(c, COLOR_HEADER_INDICE, bold=True, fuente_color=FONT_COLOR_INDICE)

    for row_idx, exc in enumerate(excepciones, 2):
        elem      = exc.get("elemento", {}) or {}
        situacion = exc.get("situacion", "")
        vals = [
            str(elem.get("ElementId", "")),
            elem.get("CodIntBIM", ""),
            elem.get(u"Categor\u00eda", elem.get("Categoria", "")),
            elem.get("Familia", ""),
            elem.get("Tipo", ""),
            elem.get("Nombre_RVT", ""),
            situacion,
        ]
        for col_idx, val in enumerate(vals, 1):
            c = ws.cell(row_idx, col_idx, value=val)
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            c.border = Border(
                left=BORDER_STYLE, right=BORDER_STYLE,
                top=BORDER_STYLE,  bottom=BORDER_STYLE
            )

    autoajustar_columnas(ws)

# ── Hoja CM (formato vertical) ────────────────────────────────
def generar_hoja_cm_vertical(ws, codigo_cm, elementos, headers_schedule, colores_dict):
    if not elementos:
        ws["A1"] = "Sin datos"
        return

    campos_extra = [
        "ElementId",
        u"Categor\u00eda",
        "Familia",
        "Tipo",
        "Nombre_RVT",
        u"Situaci\u00f3n CodIntBIM",
        "Elementos",
    ]

    # Headers base desde schedule, sin duplicar extras ni CodIntBIM
    base_headers = [
        h for h in (headers_schedule or [])
        if h and h not in campos_extra and h != "CodIntBIM"
    ]

    # Orden final de parametros (filas)
    parametros_orden = ["CodIntBIM"]
    for h in base_headers:
        if h not in parametros_orden:
            parametros_orden.append(h)
    for h in campos_extra:
        if h not in parametros_orden:
            parametros_orden.append(h)

    # Conteo por CodIntBIM (para campo "Elementos")
    conteo_por_codint = defaultdict(int)
    for elem in elementos:
        c = (elem.get("CodIntBIM") or "").strip()
        if c:
            conteo_por_codint[c] += 1

    # Columna A: nombres de parametros con color
    for row_idx, param in enumerate(parametros_orden, 1):
        c = ws.cell(row=row_idx, column=1)
        c.value = param
        aplicar_estilo_celda(c, obtener_color_parametro(param, colores_dict), bold=True)

    # Columnas B...: valores por elemento
    for col_idx, elem in enumerate(elementos, 2):
        codint        = (elem.get("CodIntBIM") or "").strip()
        mult          = conteo_por_codint.get(codint, 0)
        val_elementos = "varios" if mult > 1 else "unico"

        for row_idx, param in enumerate(parametros_orden, 1):
            c = ws.cell(row=row_idx, column=col_idx)

            if param == "ElementId":
                valor = str(elem.get("ElementId", ""))
            elif param == "CodIntBIM":
                valor = elem.get("CodIntBIM", "")
            elif param in (u"Categor\u00eda", "Categoria"):
                valor = elem.get(u"Categor\u00eda", elem.get("Categoria", ""))
            elif param == "Familia":
                valor = elem.get("Familia", "")
            elif param == "Tipo":
                valor = elem.get("Tipo", "")
            elif param == "Nombre_RVT":
                valor = elem.get("Nombre_RVT", "")
            elif param in (u"Situaci\u00f3n CodIntBIM", "Situacion CodIntBIM"):
                valor = "Elemento no Anidado"
            elif param == "Elementos":
                valor = val_elementos
            else:
                valor = elem.get(param, "")

            c.value     = valor if valor else ""
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            c.border    = Border(
                left=BORDER_STYLE, right=BORDER_STYLE,
                top=BORDER_STYLE,  bottom=BORDER_STYLE
            )

    ws.freeze_panes = "B1"
    autoajustar_columnas(ws)
    autoajustar_filas(ws)

# ── Main ──────────────────────────────────────────────────────
def main(ruta_json, ruta_xlsx_salida, data_master_dir):
    print("=" * 70)
    print("GENERACION DE EXCEL - FORMATO VERTICAL CON COLORES")
    print("=" * 70)

    datos = cargar_json(ruta_json, default={})
    if not datos:
        print("ERROR: JSON de datos vacio o no encontrado: {}".format(ruta_json))
        sys.exit(1)

    elementos_por_tabla = datos.get("elementos_por_tabla", {})
    listado_tablas      = datos.get("listado_tablas", {})
    headers_por_tabla   = datos.get("headers_por_tabla", {})
    excepciones         = datos.get("excepciones", [])

    print("  Tablas CM   : {}".format(len(elementos_por_tabla)))
    print("  Excepciones : {}".format(len(excepciones)))

    colores_dict = cargar_colores_parametros(data_master_dir)
    print("  Colores cargados: {}".format(len(colores_dict)))

    wb = Workbook()
    # Eliminar hoja por defecto
    for sh in list(wb.sheetnames):
        del wb[sh]

    # Hoja Indice
    ws_idx = wb.create_sheet("Indice", 0)
    generar_hoja_indice(ws_idx, listado_tablas)

    # Hojas CM (una por codigo)
    codigos_ordenados = sorted(elementos_por_tabla.keys())
    for i, codigo_cm in enumerate(codigos_ordenados, start=1):
        elementos        = elementos_por_tabla.get(codigo_cm, [])
        headers_schedule = headers_por_tabla.get(codigo_cm, [])
        ws_cm = wb.create_sheet(codigo_cm[:31], i)
        generar_hoja_cm_vertical(ws_cm, codigo_cm, elementos,
                                  headers_schedule, colores_dict)
        print("  Hoja '{}': {} elementos".format(codigo_cm, len(elementos)))

    # Hoja Excepciones
    if excepciones:
        ws_exc = wb.create_sheet("Excepciones", len(wb.sheetnames))
        generar_hoja_excepciones(ws_exc, excepciones)
        print("  Hoja Excepciones: {} registros".format(len(excepciones)))

    # Guardar
    directorio = os.path.dirname(ruta_xlsx_salida)
    if directorio and not os.path.exists(directorio):
        os.makedirs(directorio)

    wb.save(ruta_xlsx_salida)
    print("Excel generado: {}".format(ruta_xlsx_salida))
    print("=" * 70)

if __name__ == "__main__":
    if len(sys.argv) < 4:
        print("Uso: formatear_revision_xlsx.py <datos.json> <salida.xlsx> <data_master_dir>")
        sys.exit(1)
    main(
        sys.argv[1].strip('"').strip("'"),
        sys.argv[2].strip('"').strip("'"),
        sys.argv[3].strip('"').strip("'")
    )