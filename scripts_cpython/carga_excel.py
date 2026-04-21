# -*- coding: utf-8 -*-
"""
carga_excel.py  —  CPython 3.x
Lee la hoja CODIGO de un .xlsm desde la fila 2 (omite encabezados)
y escribe cada fila como línea de texto con valores separados por ";"
en el archivo temporal indicado como segundo argumento.

Uso:
    python carga_excel.py <ruta_xlsm> <ruta_tmp_txt>
"""
import os
import sys


def read_codigos_sheet(xlsm_path):
    try:
        from openpyxl import load_workbook
    except ImportError:
        print(u"ERROR: openpyxl no está instalado. Ejecuta: pip install openpyxl")
        return None

    if not os.path.isfile(xlsm_path):
        print(u"ERROR: No existe el archivo: {}".format(xlsm_path))
        return None

    try:
        wb = load_workbook(xlsm_path, data_only=True, read_only=True)
    except Exception as e:
        print(u"ERROR abriendo Excel: {}".format(e))
        return None

    if "CODIGO" not in wb.sheetnames:
        print(u"ERROR: No se encontró la hoja 'CODIGO' en el archivo.")
        wb.close()
        return None

    filas = []
    try:
        sheet = wb["CODIGO"]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if any(cell is not None for cell in row):
                filas.append(row)
    except Exception as e:
        print(u"ERROR leyendo hoja CODIGO: {}".format(e))
        wb.close()
        return None

    wb.close()
    return filas


def main():
    if len(sys.argv) < 3:
        print(u"Uso: carga_excel.py <ruta_xlsm> <ruta_tmp_txt>")
        sys.exit(1)

    xlsm_path    = sys.argv[1]
    tmp_txt_path = sys.argv[2]

    filas = read_codigos_sheet(xlsm_path)
    if filas is None:
        sys.exit(1)

    if not filas:
        print(u"AVISO: La hoja CODIGO no tiene datos a partir de la fila 2.")
        # Crear archivo vacío para que script.py no falle
        try:
            tmp_dir = os.path.dirname(tmp_txt_path)
            if tmp_dir and not os.path.exists(tmp_dir):
                os.makedirs(tmp_dir)
            with open(tmp_txt_path, "w", encoding="utf-8") as f:
                f.write("")
        except Exception as e:
            print(u"ERROR creando archivo temporal vacío: {}".format(e))
            sys.exit(1)
        sys.exit(0)

    try:
        tmp_dir = os.path.dirname(tmp_txt_path)
        if tmp_dir and not os.path.exists(tmp_dir):
            os.makedirs(tmp_dir)
        with open(tmp_txt_path, "w", encoding="utf-8") as f:
            for row in filas:
                valores = [u"" if v is None else str(v).strip() for v in row]
                f.write(u";".join(valores) + u"\n")
    except Exception as e:
        print(u"ERROR escribiendo archivo temporal: {}".format(e))
        sys.exit(1)

    print(u"OK: {} filas escritas en {}".format(len(filas), tmp_txt_path))
    sys.exit(0)


if __name__ == "__main__":
    main()