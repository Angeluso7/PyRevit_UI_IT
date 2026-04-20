# -*- coding: utf-8 -*-
"""
formatear_tablas_planilla_vs_modelo.py  v1.0

Formato HORIZONTAL con colores por estado de comparación:
- Fila 1: headers (nombres de parámetros)
- Filas siguientes: una fila por CodIntBIM comparado
- Colores por estado de celda:
    ok           → #C6EFCE  verde claro
    falta_modelo → #FFC7CE  rojo claro
    falta_excel  → #FFEB9C  amarillo
    difiere      → #F4B084  naranja
    no_existe    → #DAEEF3  celeste gris

Estructura del JSON de entrada (comparacion.json generado por ui_comparacion.py):
{
  "listado_tablas":    {"valores": [...], "claves": [...]},
  "elementos_por_tabla": { "CMxx": [ {header: valor, ...}, ... ] },
  "headers_por_tabla":   { "CMxx": [...] }
}

Cada elemento en elementos_por_tabla ya tiene los valores fusionados
(campo único con "Planilla: X\\nModelo: Y" si difieren).
Para derivar el estado de cada celda se re-evalúa el valor:
  - vacío                          → no_existe
  - empieza con "Planilla:"        → difiere
  - empieza con "Modelo:"          → falta_excel  (solo en modelo)
  - empieza con "Falta en Modelo"  → falta_modelo
  Cualquier otro valor no vacío    → ok

Uso:
    python formatear_tablas_planilla_vs_modelo.py <ruta_comparacion.json> <ruta_xlsx_salida>
"""

import sys
import os
import json
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Colores de estado ─────────────────────────────────────────────────────────
COLOR_OK            = "FFC6EFCE"   # verde claro
COLOR_FALTA_MODELO  = "FFFFC7CE"   # rojo claro
COLOR_FALTA_EXCEL   = "FFFFEB9C"   # amarillo
COLOR_DIFIERE       = "FFF4B084"   # naranja
COLOR_NO_EXISTE     = "FFDAEEF3"   # celeste gris
COLOR_HEADER_CM     = "FF366092"   # azul oscuro — cabecera de tabla CM
COLOR_HEADER_IDX    = "FF1F3864"   # azul marino — cabecera de índice
COLOR_CODINTBIM     = "FFD7D7D7"   # gris — columna CodIntBIM
FONT_INV            = "FFFFFFFF"   # blanco — fuente sobre fondos oscuros

BORDER = Side(style="thin", color="FF000000")
FULL_BORDER = Border(left=BORDER, right=BORDER, top=BORDER, bottom=BORDER)


# ── Helpers de estilo ─────────────────────────────────────────────────────────

def _fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _font(bold=False, color="FF000000", size=10):
    return Font(bold=bold, color=color, size=size)


def _alinear(wrap=True, horizontal="left"):
    return Alignment(horizontal=horizontal, vertical="center", wrap_text=wrap)


def _aplicar(cell, color_hex, bold=False, font_color="FF000000",
             wrap=True, horizontal="left"):
    cell.fill      = _fill(color_hex)
    cell.font      = _font(bold=bold, color=font_color)
    cell.alignment = _alinear(wrap=wrap, horizontal=horizontal)
    cell.border    = FULL_BORDER


def _autoajustar_columnas(ws, min_w=12, max_w=55):
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    # ignorar saltos de línea para el cálculo de ancho
                    lineas = str(cell.value).split("\n")
                    max_len = max(max_len, max(len(l) for l in lineas))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_w), max_w)


def _autoajustar_filas(ws, altura=18):
    for row_idx in range(1, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = altura


# ── Inferir estado desde valor fusionado ─────────────────────────────────────

def _inferir_estado(valor):
    """
    Re-evalúa el estado de comparación a partir del texto del valor fusionado.
    ui_comparacion.py escribe:
        "Planilla: X\\nModelo: Y"   → difiere
        "Planilla: X\\n"            → falta_modelo  (solo en excel)
        "Modelo: X\\nPlanilla: "    → falta_excel   (solo en modelo)
        valor normal                → ok
        vacío                       → no_existe
    """
    v = (valor or "").strip()
    if not v:
        return "no_existe"
    if v.startswith("Planilla:") and "Modelo:" in v:
        # tiene ambos → difieren
        return "difiere"
    if v.startswith("Planilla:"):
        # solo planilla → falta en modelo
        return "falta_modelo"
    if v.startswith("Modelo:"):
        # solo modelo → falta en excel
        return "falta_excel"
    return "ok"


def _color_estado(estado):
    mapa = {
        "ok":           COLOR_OK,
        "falta_modelo": COLOR_FALTA_MODELO,
        "falta_excel":  COLOR_FALTA_EXCEL,
        "difiere":      COLOR_DIFIERE,
        "no_existe":    COLOR_NO_EXISTE,
    }
    return mapa.get(estado, COLOR_NO_EXISTE)


# ── Hoja Índice ───────────────────────────────────────────────────────────────

def _generar_hoja_indice(ws, listado_tablas):
    ws.title = "Índice"

    # Título
    ws["B2"] = "Planilla vs Modelo — Listado de Tablas"
    ws["B2"].font      = _font(bold=True, size=12, color="FF000000")
    ws["B2"].alignment = _alinear(wrap=False)

    # Cabecera
    ws["B4"] = "Código CM"
    ws["C4"] = "Nombre Planilla"
    ws["D4"] = "Elementos"
    for col in ("B", "C", "D"):
        _aplicar(ws["{}4".format(col)], COLOR_HEADER_IDX,
                 bold=True, font_color=FONT_INV, wrap=False)

    valores = listado_tablas.get("valores", [])   # códigos CM
    claves  = listado_tablas.get("claves",  [])   # nombres de planilla

    pares = sorted(zip(valores, claves), key=lambda x: x[0])
    for idx, (codigo, nombre) in enumerate(pares):
        fila = 5 + idx
        ws["B{}".format(fila)] = codigo
        ws["C{}".format(fila)] = nombre
        ws["D{}".format(fila)] = ""   # se rellena al final si se quiere
        for col in ("B", "C", "D"):
            cell = ws["{}{}".format(col, fila)]
            cell.alignment = _alinear(wrap=False)
            cell.border    = FULL_BORDER

    _autoajustar_columnas(ws)


# ── Hoja Leyenda ──────────────────────────────────────────────────────────────

def _generar_hoja_leyenda(ws):
    ws.title = "Leyenda"

    ws["B2"] = "Leyenda de colores — Planilla vs Modelo"
    ws["B2"].font      = _font(bold=True, size=12)
    ws["B2"].alignment = _alinear(wrap=False)

    estados = [
        ("ok",           COLOR_OK,           "OK / Coincide",
         "El valor en Planilla y en Modelo es idéntico."),
        ("falta_modelo", COLOR_FALTA_MODELO,  "Falta en Modelo",
         "El valor existe en la Planilla pero no está en el Modelo Revit."),
        ("falta_excel",  COLOR_FALTA_EXCEL,   "Falta en Excel",
         "El valor existe en el Modelo Revit pero no está en la Planilla."),
        ("difiere",      COLOR_DIFIERE,       "Diferencia de valor",
         "Ambos tienen valor pero no coinciden. La celda muestra 'Planilla: X / Modelo: Y'."),
        ("no_existe",    COLOR_NO_EXISTE,     "No existe",
         "El campo está vacío en ambas fuentes."),
    ]

    ws["B4"] = "Color"
    ws["C4"] = "Estado"
    ws["D4"] = "Descripción"
    for col in ("B", "C", "D"):
        _aplicar(ws["{}4".format(col)], COLOR_HEADER_IDX,
                 bold=True, font_color=FONT_INV, wrap=False)

    for idx, (estado, color, nombre, desc) in enumerate(estados):
        fila = 5 + idx
        _aplicar(ws["B{}".format(fila)], color, wrap=False, horizontal="center")
        ws["C{}".format(fila)].value     = nombre
        ws["C{}".format(fila)].border    = FULL_BORDER
        ws["C{}".format(fila)].alignment = _alinear(wrap=False)
        ws["C{}".format(fila)].font      = _font(bold=True)
        ws["D{}".format(fila)].value     = desc
        ws["D{}".format(fila)].border    = FULL_BORDER
        ws["D{}".format(fila)].alignment = _alinear(wrap=True)

    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 70


# ── Hoja CM (tabla comparativa horizontal) ────────────────────────────────────

def _generar_hoja_cm(ws, codigo_cm, elementos, headers):
    """
    Formato HORIZONTAL:
      Fila 1  : cabecera con nombres de parámetros
      Filas 2+: una fila por elemento/CodIntBIM
    Colores por celda según estado inferido del valor.
    """
    if not elementos:
        ws["A1"] = "Sin datos"
        ws["A1"].font = _font(bold=True)
        return

    if not headers:
        # reconstruir headers desde claves del primer elemento
        claves = list(elementos[0].keys())
        headers = (["CodIntBIM"] +
                   [k for k in claves if k not in ("CodIntBIM", "estado_por_celda")])

    # ── Fila 1: cabecera ─────────────────────────────────────────────────────
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = h
        if h == "CodIntBIM":
            _aplicar(cell, COLOR_CODINTBIM, bold=True, wrap=False)
        else:
            _aplicar(cell, COLOR_HEADER_CM, bold=True,
                     font_color=FONT_INV, wrap=False)

    # ── Filas de datos ────────────────────────────────────────────────────────
    for row_idx, elem in enumerate(elementos, 2):
        # Si el elemento ya trae lista de estados (generado por ui_comparacion)
        estados_precalc = elem.get("estado_por_celda") or []

        for col_idx, h in enumerate(headers, 1):
            cell  = ws.cell(row=row_idx, column=col_idx)
            valor = elem.get(h, "") or ""

            # Determinar estado
            if h == "CodIntBIM":
                estado = "ok" if valor else "no_existe"
                color  = COLOR_CODINTBIM
            else:
                if col_idx - 1 < len(estados_precalc):
                    estado = estados_precalc[col_idx - 1]
                else:
                    estado = _inferir_estado(valor)
                color = _color_estado(estado)

            cell.value = valor if valor else ""
            _aplicar(cell, color, wrap=True)

    # Congelar primera fila (cabecera) y primera columna (CodIntBIM)
    ws.freeze_panes = "B2"
    _autoajustar_columnas(ws)
    _autoajustar_filas(ws)


# ── Hoja Resumen estadístico ──────────────────────────────────────────────────

def _generar_hoja_resumen(ws, elementos_por_tabla, headers_por_tabla):
    """
    Una fila por tabla CM con conteo de estados:
    OK | Falta Modelo | Falta Excel | Difiere | No Existe | Total
    """
    ws.title = "Resumen"

    ws["B2"] = "Resumen de Comparación por Tabla"
    ws["B2"].font      = _font(bold=True, size=12)
    ws["B2"].alignment = _alinear(wrap=False)

    cabeceras = ["Código CM", "Total Elementos",
                 "OK", "Falta Modelo", "Falta Excel", "Difiere", "No Existe"]

    for col_idx, cab in enumerate(cabeceras, 2):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = cab
        _aplicar(cell, COLOR_HEADER_IDX, bold=True, font_color=FONT_INV,
                 wrap=False, horizontal="center")

    row_idx = 5
    totales = {"ok": 0, "falta_modelo": 0, "falta_excel": 0,
               "difiere": 0, "no_existe": 0}

    for codigo_cm in sorted(elementos_por_tabla.keys()):
        elementos = elementos_por_tabla.get(codigo_cm, [])
        headers   = headers_por_tabla.get(codigo_cm, [])
        conteos   = {"ok": 0, "falta_modelo": 0, "falta_excel": 0,
                     "difiere": 0, "no_existe": 0}

        for elem in elementos:
            estados_precalc = elem.get("estado_por_celda") or []
            for col_pos, h in enumerate(headers):
                if h == "CodIntBIM":
                    continue
                if col_pos < len(estados_precalc):
                    estado = estados_precalc[col_pos]
                else:
                    estado = _inferir_estado(elem.get(h, ""))
                if estado in conteos:
                    conteos[estado] += 1

        total_celdas = sum(conteos.values())
        for k in totales:
            totales[k] += conteos[k]

        valores_fila = [
            codigo_cm,
            len(elementos),
            conteos["ok"],
            conteos["falta_modelo"],
            conteos["falta_excel"],
            conteos["difiere"],
            conteos["no_existe"],
        ]
        colores_fila = [
            COLOR_CODINTBIM,  # Código CM
            "FFFFFFFF",       # Total Elementos
            COLOR_OK,
            COLOR_FALTA_MODELO,
            COLOR_FALTA_EXCEL,
            COLOR_DIFIERE,
            COLOR_NO_EXISTE,
        ]

        for col_idx, (val, color) in enumerate(zip(valores_fila, colores_fila), 2):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = val
            _aplicar(cell, color, wrap=False,
                     horizontal="center" if col_idx > 3 else "left")

        row_idx += 1

    # Fila de totales
    ws.cell(row=row_idx, column=2).value = "TOTAL"
    ws.cell(row=row_idx, column=2).font  = _font(bold=True)
    ws.cell(row=row_idx, column=2).border = FULL_BORDER

    totales_fila = [
        "",
        totales["ok"],
        totales["falta_modelo"],
        totales["falta_excel"],
        totales["difiere"],
        totales["no_existe"],
    ]
    colores_totales = [
        "FFFFFFFF",
        COLOR_OK,
        COLOR_FALTA_MODELO,
        COLOR_FALTA_EXCEL,
        COLOR_DIFIERE,
        COLOR_NO_EXISTE,
    ]
    for col_idx, (val, color) in enumerate(zip(totales_fila, colores_totales), 3):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = val
        _aplicar(cell, color, bold=True, wrap=False, horizontal="center")

    _autoajustar_columnas(ws)
    _autoajustar_filas(ws, altura=20)


# ── MAIN ──────────────────────────────────────────────────────────────────────

def main(ruta_json, ruta_xlsx_salida):
    print("=" * 70)
    print("PLANILLA VS MODELO — GENERACIÓN EXCEL COMPARATIVO")
    print("=" * 70)

    if not os.path.exists(ruta_json):
        print("ERROR: no se encontró el JSON de comparación -> {}".format(ruta_json))
        sys.exit(1)

    with open(ruta_json, "r", encoding="utf-8") as f:
        datos = json.load(f)

    elementos_por_tabla = datos.get("elementos_por_tabla", {})
    listado_tablas      = datos.get("listado_tablas",      {})
    headers_por_tabla   = datos.get("headers_por_tabla",   {})

    print(" - Tablas CM encontradas : {}".format(len(elementos_por_tabla)))
    total_elems = sum(len(v) for v in elementos_por_tabla.values())
    print(" - Total de elementos    : {}".format(total_elems))

    wb = Workbook()
    # eliminar hoja por defecto
    if wb.sheetnames:
        wb.remove(wb.active)

    # 1. Índice
    ws_idx = wb.create_sheet("Índice", 0)
    _generar_hoja_indice(ws_idx, listado_tablas)
    print(" ✓ Hoja 'Índice' generada")

    # 2. Resumen
    ws_res = wb.create_sheet("Resumen", 1)
    _generar_hoja_resumen(ws_res, elementos_por_tabla, headers_por_tabla)
    print(" ✓ Hoja 'Resumen' generada")

    # 3. Leyenda
    ws_ley = wb.create_sheet("Leyenda", 2)
    _generar_hoja_leyenda(ws_ley)
    print(" ✓ Hoja 'Leyenda' generada")

    # 4. Una hoja por cada CM
    codigos_ordenados = sorted(elementos_por_tabla.keys())
    for i, codigo_cm in enumerate(codigos_ordenados, start=3):
        elementos = elementos_por_tabla.get(codigo_cm, [])
        headers   = headers_por_tabla.get(codigo_cm,   [])
        nombre_hoja = codigo_cm[:31]
        ws_cm = wb.create_sheet(nombre_hoja, i)
        _generar_hoja_cm(ws_cm, codigo_cm, elementos, headers)
        print(" ✓ Hoja '{}' — {} elementos".format(nombre_hoja, len(elementos)))

    # Guardar
    try:
        wb.save(ruta_xlsx_salida)
        print("")
        print("✓ Archivo generado correctamente:")
        print("  {}".format(ruta_xlsx_salida))
    except Exception as e:
        print("ERROR al guardar el archivo: {}".format(e))
        sys.exit(1)

    print("=" * 70)


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Uso: formatear_tablas_planilla_vs_modelo.py "
              "<ruta_comparacion.json> <ruta_xlsx_salida>")
        sys.exit(1)
    main(sys.argv[1], sys.argv[2])