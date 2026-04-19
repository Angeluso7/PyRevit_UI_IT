# -*- coding: utf-8 -*-
"""
exportar_planillas_xlsx.py  (CPython 3.x)
==========================================
Punto de entrada CPython para el boton "Por Seleccion".

Uso:
    python exportar_planillas_xlsx.py <csv_path> [script_json_path]

Lee el CSV generado por Revit (delimitador ';', encoding utf-8),
construye la estructura de datos esperada y aplica el formato VERTICAL
con colores por parametro definidos en colores_parametros.json.

Salida:  <mismo nombre del CSV pero con extension .xlsx>
"""
import sys
import os
import json
import csv

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as e:
    sys.exit("ERROR: openpyxl no instalado.\n{}".format(e))

# ── constantes de color ────────────────────────────────────────────────────
COLOR_DEFAULT_NO_JSON = "FFFFCC"
COLOR_ADICIONALES     = "D7D7D7"
COLOR_CODINTBIM       = "D7D7D7"
COLOR_HEADER_INDICE   = "366092"
FONT_COLOR_INDICE     = "FFFFFF"

PARAMETROS_ADICIONALES = {
    "ElementId", "Categoria", "Familia", "Tipo",
    "Nombre_RVT", "Situacion CodIntBIM", "Elementos",
}

BORDER_SIDE  = Side(style="thin", color="FF000000")
FULL_BORDER  = Border(
    left=BORDER_SIDE, right=BORDER_SIDE,
    top=BORDER_SIDE,  bottom=BORDER_SIDE
)

# ── helpers ────────────────────────────────────────────────────────────────
def _load_json(path):
    if not os.path.isfile(path):
        return {}
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as exc:
        sys.stderr.write("AVISO: no se pudo leer {}: {}\n".format(path, exc))
        return {}

def _resolve_data_dir(script_json_path):
    if script_json_path and os.path.isfile(script_json_path):
        return os.path.dirname(script_json_path)
    return os.path.join(
        os.path.expanduser("~"),
        r"AppData\Roaming\MyPyRevitExtention\PyRevitIT.extension\data"
    )

def _fill(hex6):
    c = hex6.lstrip("#").lstrip("FF")
    c = c if len(c) == 6 else "D7D7D7"
    return PatternFill(start_color=c, end_color=c, fill_type="solid")

def _estilo(cell, color6, bold=False, font_color="000000"):
    cell.fill      = _fill(color6)
    cell.font      = Font(bold=bold, color=font_color)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border    = FULL_BORDER

def _color_param(param, colores):
    if param == "CodIntBIM":
        return COLOR_CODINTBIM
    if param in PARAMETROS_ADICIONALES:
        return COLOR_ADICIONALES
    raw = colores.get(param, "").strip().lstrip("#")
    return raw if len(raw) in (6, 8) else COLOR_DEFAULT_NO_JSON

def _autoajustar(ws):
    for ci in range(1, ws.max_column + 1):
        ml = 0
        for ri in range(1, ws.max_row + 1):
            v = ws.cell(ri, ci).value
            if v:
                ml = max(ml, len(str(v)))
        ws.column_dimensions[get_column_letter(ci)].width = min(max(ml + 2, 15), 60)
    for ri in range(1, ws.max_row + 1):
        ws.row_dimensions[ri].height = 20

# ── hojas ──────────────────────────────────────────────────────────────────
def _hoja_indice(ws, nombre_schedule):
    ws.title = "Indice"
    ws["B3"] = "Listado de Tablas"
    ws["B3"].font = Font(bold=True, size=12)
    ws["B4"] = "Nombre Schedule"
    _estilo(ws["B4"], COLOR_HEADER_INDICE, bold=True, font_color=FONT_COLOR_INDICE)
    ws["B5"] = nombre_schedule
    _autoajustar(ws)

def _orden_param(h):
    if h == "CodIntBIM":        return (0, h)
    if h in PARAMETROS_ADICIONALES: return (2, h)
    return (1, h)

def _hoja_vertical(ws, titulo, headers, filas, colores):
    ws.title = titulo[:31]
    if not filas:
        ws["A1"] = "Sin datos"
        return

    # Encabezado columnas
    ws.cell(1, 1, value="Parametro")
    _estilo(ws.cell(1, 1), COLOR_HEADER_INDICE, bold=True, font_color=FONT_COLOR_INDICE)
    for idx in range(len(filas)):
        c = ws.cell(1, idx + 2)
        c.value = "Elemento {}".format(idx + 1)
        _estilo(c, COLOR_HEADER_INDICE, bold=True, font_color=FONT_COLOR_INDICE)

    headers_ord = sorted(headers, key=_orden_param)

    for ri, param in enumerate(headers_ord, start=2):
        c_param = ws.cell(ri, 1)
        c_param.value = param
        color = _color_param(param, colores)
        _estilo(c_param, color, bold=(param == "CodIntBIM"))

        pidx = headers.index(param) if param in headers else None
        for ci, fila in enumerate(filas, start=2):
            c_val = ws.cell(ri, ci)
            c_val.value = fila[pidx] if (pidx is not None and pidx < len(fila)) else ""
            _estilo(c_val, color)

    ws.freeze_panes = "B2"
    _autoajustar(ws)

# ── lectura CSV ────────────────────────────────────────────────────────────
def leer_csv(csv_path):
    rows = []
    with open(csv_path, "r", encoding="utf-8-sig") as f:
        for row in csv.reader(f, delimiter=";"):
            if any(c.strip() for c in row):
                rows.append(row)
    if not rows:
        return [], []
    return [h.strip() for h in rows[0]], rows[1:]

# ── main ──────────────────────────────────────────────────────────────────
def main():
    if len(sys.argv) < 2:
        sys.exit("Uso: python exportar_planillas_xlsx.py <csv_path> [script_json_path]")

    csv_path        = sys.argv[1]
    json_path       = sys.argv[2] if len(sys.argv) > 2 else None

    if not os.path.isfile(csv_path):
        sys.exit("ERROR: no se encontro el CSV: {}".format(csv_path))

    data_dir    = _resolve_data_dir(json_path)
    master_dir  = os.path.join(data_dir, "master")

    script_json  = _load_json(json_path) if json_path else {}
    colores_path = os.path.join(master_dir, "colores_parametros.json")
    if not os.path.isfile(colores_path):
        colores_path = os.path.join(data_dir, "colores_parametros.json")
    colores = _load_json(colores_path)

    reemplazos = script_json.get("reemplazos_encabezados", {})
    nombre     = os.path.splitext(os.path.basename(csv_path))[0]
    headers_raw, filas = leer_csv(csv_path)

    if not headers_raw:
        sys.exit("ERROR: CSV vacio o sin encabezados: {}".format(csv_path))

    headers = [reemplazos.get(h, h) for h in headers_raw]

    wb        = Workbook()
    ws_indice = wb.active
    _hoja_indice(ws_indice, nombre)

    ws_datos = wb.create_sheet()
    _hoja_vertical(ws_datos, nombre[:31], headers, filas, colores)

    xlsx_path = os.path.splitext(csv_path)[0] + ".xlsx"
    wb.save(xlsx_path)
    print("OK:{}".format(xlsx_path))

if __name__ == "__main__":
    main()
