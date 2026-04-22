# -*- coding: utf-8 -*-
"""
formatear_tablas_excel.py — CPython 3
Lee comparacion.json y genera un .xlsx con:
  - Hoja Índice
  - Hojas por código CMxx con tabla comparativa y leyenda de colores

Uso:
    python formatear_tablas_excel.py <ruta_comparacion_json> <ruta_xlsx_salida>

Argumentos:
    sys.argv[1] = ruta al comparacion.json (generado por ui_comparacion.py)
    sys.argv[2] = ruta destino del archivo .xlsx
"""

import sys
import os
import json
import traceback
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

COLOR_ESTADOS = {
    "ok"           : "FFC6EFCE",   # verde claro
    "falta_modelo" : "FFFFC7CE",   # rojo claro
    "falta_excel"  : "FFFFEB9C",   # amarillo
    "difiere"      : "FFF4B084",   # naranja
    "no_existe"    : "FFD9D9D9",   # gris claro
}
HEADER_COLOR     = "FF366092"
HEADER_FONT_CLR  = "FFFFFFFF"
BORDER_STYLE     = Side(style="thin", color="FF000000")


# ── Helpers de formato ────────────────────────────────────────
def _border():
    return Border(
        left=BORDER_STYLE, right=BORDER_STYLE,
        top=BORDER_STYLE,  bottom=BORDER_STYLE
    )

def aplicar_encabezado(ws, row_num, n_cols):
    hfont = Font(bold=True, color=HEADER_FONT_CLR)
    hfill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
    halig = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col in range(1, n_cols + 1):
        c = ws.cell(row=row_num, column=col)
        c.font      = hfont
        c.fill      = hfill
        c.alignment = halig
        c.border    = _border()

def auto_ajustar_columnas(ws, max_ancho=60):
    for col in range(1, ws.max_column + 1):
        max_len = 0
        letra   = get_column_letter(col)
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[letra].width = min(max_len + 2, max_ancho)

def agregar_leyenda(ws, fila_inicio):
    items = [
        ("OK / Coincide",     "ok"),
        ("Falta en modelo",   "falta_modelo"),
        ("Falta en Excel",    "falta_excel"),
        ("Diferencia",        "difiere"),
        ("No existe",         "no_existe"),
    ]
    for i, (texto, estado) in enumerate(items):
        color = COLOR_ESTADOS.get(estado, "FFFFFFFF")
        ws.cell(row=fila_inicio + i, column=1).fill = PatternFill(
            start_color=color, end_color=color, fill_type="solid"
        )
        ws.cell(row=fila_inicio + i, column=2).value = texto


# ── Hoja Índice ───────────────────────────────────────────────
def generar_hoja_indice(ws, listado_tablas):
    ws.title = "Índice"
    ws["B3"]      = "Listado de Tablas"
    ws["B3"].font = Font(bold=True, size=12, color=HEADER_COLOR)

    ws["B4"] = "Código"
    ws["C4"] = "Nombre Tabla (Schedule)"
    aplicar_encabezado(ws, 4, 2)
    # Encabezados están en columnas B y C → reescribir valores
    ws["B4"] = "Código"
    ws["C4"] = "Nombre Tabla (Schedule)"

    valores = listado_tablas.get("valores", [])
    claves  = listado_tablas.get("claves",  [])
    pares   = sorted(zip(valores, claves), key=lambda x: x[0])

    for idx, (codigo, nombre) in enumerate(pares):
        fila = 5 + idx
        ws.cell(row=fila, column=2).value = codigo
        ws.cell(row=fila, column=3).value = nombre

    auto_ajustar_columnas(ws)


# ── Hoja por CM ───────────────────────────────────────────────
def generar_hoja_tabla(ws, bloque):
    headers = bloque.get("headers", [])
    filas   = bloque.get("filas",   [])
    ws.freeze_panes = "A2"

    if not headers:
        ws.cell(row=1, column=1).value = "Sin datos"
        return

    # Fila 1: encabezados
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=h)
    aplicar_encabezado(ws, 1, len(headers))
    # Restaurar valores tras aplicar encabezado
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx).value = h

    # Filas de datos
    fila_actual = 2
    for fila in filas:
        valores = fila.get("valores", [])
        estados = fila.get("estado_por_celda", [])

        for col_idx in range(1, len(headers) + 1):
            val    = valores[col_idx - 1] if col_idx - 1 < len(valores) else ""
            estado = estados[col_idx - 1] if col_idx - 1 < len(estados) else "ok"
            color  = COLOR_ESTADOS.get(estado, COLOR_ESTADOS["ok"])

            c = ws.cell(row=fila_actual, column=col_idx, value=val)
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            c.border    = _border()
            c.fill      = PatternFill(start_color=color, end_color=color, fill_type="solid")

        fila_actual += 1

    # Leyenda
    fila_leyenda = fila_actual + 2
    ws.cell(row=fila_leyenda, column=1).value = "Leyenda:"
    ws.cell(row=fila_leyenda, column=1).font  = Font(bold=True)
    agregar_leyenda(ws, fila_leyenda + 1)

    auto_ajustar_columnas(ws)


# ── Main ──────────────────────────────────────────────────────
def main(ruta_json, ruta_xlsx):
    # Validar que el JSON existe
    if not os.path.exists(ruta_json):
        raise FileNotFoundError(
            "comparacion.json no encontrado: {}".format(ruta_json)
        )

    with open(ruta_json, "r", encoding="utf-8") as f:
        datos = json.load(f)

    listado_tablas  = datos.get("listado_tablas", {})
    datos_por_tabla = datos.get("datos_por_tabla", {})

    if not datos_por_tabla:
        raise ValueError(
            "El JSON no contiene 'datos_por_tabla' o está vacío.\n"
            "Claves presentes: {}".format(list(datos.keys()))
        )

    # Crear directorio de salida si no existe
    directorio_xlsx = os.path.dirname(ruta_xlsx)
    if directorio_xlsx and not os.path.exists(directorio_xlsx):
        os.makedirs(directorio_xlsx)

    wb = Workbook()
    # Eliminar hoja por defecto
    for sh in wb.sheetnames:
        del wb[sh]

    # Hoja Índice
    ws_idx = wb.create_sheet("Índice", 0)
    generar_hoja_indice(ws_idx, listado_tablas)

    # Hojas por CM (ordenadas)
    for sheet_idx, codigo in enumerate(sorted(datos_por_tabla.keys()), 1):
        bloque     = datos_por_tabla[codigo]
        nombre_hoja = codigo[:31]
        ws          = wb.create_sheet(nombre_hoja, sheet_idx)
        generar_hoja_tabla(ws, bloque)

    wb.save(ruta_xlsx)
    print("Archivo generado: {}".format(ruta_xlsx))


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print(
            "Uso: formatear_tablas_excel.py <comparacion_json> <ruta_xlsx_salida>",
            file=sys.stderr
        )
        sys.exit(1)

    ruta_json = sys.argv[1].strip('"').strip("'")
    ruta_xlsx = sys.argv[2].strip('"').strip("'")

    try:
        main(ruta_json, ruta_xlsx)
    except Exception as e:
        # Imprimir error COMPLETO en stderr para que CalledProcessError lo muestre
        print("=" * 60, file=sys.stderr)
        print("ERROR en formatear_tablas_excel.py:", file=sys.stderr)
        print(str(e), file=sys.stderr)
        print("-" * 60, file=sys.stderr)
        traceback.print_exc(file=sys.stderr)
        print("=" * 60, file=sys.stderr)
        sys.exit(1)