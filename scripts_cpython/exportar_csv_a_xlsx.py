# -*- coding: utf-8 -*-
import os
import sys
import csv
import json
from datetime import datetime
from copy import copy
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

ENCABEZADOS_GRIS = [
    "val-sql", "val-activos", "id", "nombre", "estado", "Nodo",
    "Propietario", "Decreto", "Categoría VU", "Estructuras", "OOCC",
    "Armario", "DEA", "Línea", "Subestación", "Comuna", "Patio",
    "Antecedentes complementarios", "Sala SSGG y/o Caseta", "Tramo",
    "Nodo 1", "Nodo 2", "Nodo 3", "Circuito", "SS.GG., Caseta y Salas de ER",
    "Sistema Eléctrico", "Nodo Paño/Barra", "Nodo SSEE", "Paño", "Servidumbres",
    "Vano", "Nodo SS.GG. y Caseta", "Nodo Paño", "Nodo de Estructura",
    "inf-id"
]

COLOR_GRIS = "808080"
COLOR_AZUL = "0C769E"
FONT_ENCABEZADO = Font(name='Calibri', bold=True, color='FFFFFF')
FONT_TITULO = Font(name='Calibri', size=14, bold=True)
FONT_INFO = Font(name='Calibri', size=11)
ALIGN_CENTRO = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_IZQ = Alignment(horizontal='left', vertical='center')
BORDE_DELGADO = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
RELLENO_GRIS = PatternFill('solid', fgColor=COLOR_GRIS)
RELLENO_AZUL = PatternFill('solid', fgColor=COLOR_AZUL)


def get_script_json_path(cli_json_path=None):
    if cli_json_path:
        return cli_json_path
    return os.path.join(os.path.expanduser('~'), r'AppData\Roaming\MyPyRevitExtention\PyRevitIT.extension\data\script.json')


def cargar_config_json(json_path=None):
    json_path = get_script_json_path(json_path)
    if not os.path.exists(json_path):
        return {}, {}
    with open(json_path, 'r', encoding='utf-8') as f:
        config = json.load(f)
    return config.get('reemplazos_encabezados', {}), config.get('reemplazos_de_nombres', {})


def convertir_a_string(valor):
    if isinstance(valor, tuple):
        return ' '.join(convertir_a_string(v) for v in valor)
    if valor is None:
        return ''
    try:
        return str(valor)
    except Exception:
        return ''


def crear_excel_base(csv_path, xlsx_path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Registros'

    with open(csv_path, newline='', encoding='utf-8-sig') as f:
        reader = csv.reader(f, delimiter=';')
        filas = [row for row in reader if row]

    n_encabezado_filas = 5
    n_cols = max(len(row) for row in filas) if filas else 0

    for r_idx, fila in enumerate(filas, start=n_encabezado_filas + 1):
        for c_idx, valor in enumerate(fila, start=1):
            celda = ws.cell(row=r_idx, column=c_idx)
            valor_str = convertir_a_string(valor)
            celda.value = valor_str
            if r_idx == n_encabezado_filas + 1:
                celda.fill = RELLENO_GRIS if valor_str in ENCABEZADOS_GRIS else RELLENO_AZUL
                celda.font = FONT_ENCABEZADO
                celda.alignment = ALIGN_CENTRO
                celda.border = BORDE_DELGADO
            else:
                celda.alignment = ALIGN_IZQ
                celda.border = BORDE_DELGADO

    for col in range(1, n_cols + 1):
        letra_col = ws.cell(row=n_encabezado_filas + 1, column=col).column_letter
        ws.column_dimensions[letra_col].width = 15

    wb.save(xlsx_path)


def copiar_fila_formato(ws, fila_origen, fila_destino, max_col):
    for col in range(1, max_col + 1):
        origen = ws.cell(row=fila_origen, column=col)
        destino = ws.cell(row=fila_destino, column=col)
        destino.font = copy(origen.font)
        destino.border = copy(origen.border)
        destino.fill = copy(origen.fill)
        destino.alignment = copy(origen.alignment)
        destino.number_format = copy(origen.number_format)
        destino.protection = copy(origen.protection)


def procesar_encabezados_divididos(ws, fila_encabezado, reemplazos_encabezados):
    max_col = ws.max_column
    max_segmentos = 1
    titulos_segmentados = []

    for col in range(1, max_col + 1):
        celda = ws.cell(row=fila_encabezado, column=col)
        val = celda.value if celda.value else ''
        if val in reemplazos_encabezados:
            val = reemplazos_encabezados[val]
        segmentos = [seg.strip() for seg in val.split('/')]
        titulos_segmentados.append(segmentos)
        if len(segmentos) > max_segmentos:
            max_segmentos = len(segmentos)

    if max_segmentos > 1:
        ws.insert_rows(fila_encabezado + 1, max_segmentos - 1)
        for i in range(1, max_segmentos):
            copiar_fila_formato(ws, fila_encabezado, fila_encabezado + i, max_col)

    for col in range(1, max_col + 1):
        segmentos = titulos_segmentados[col - 1]
        for i, texto in enumerate(segmentos):
            fila = fila_encabezado + i
            celda = ws.cell(row=fila, column=col)
            celda.value = texto
            celda.font = copy(ws.cell(row=fila_encabezado, column=col).font)
            celda.alignment = ALIGN_CENTRO
            celda.border = copy(ws.cell(row=fila_encabezado, column=col).border)
            celda.fill = copy(ws.cell(row=fila_encabezado, column=col).fill)

    for nivel in range(max_segmentos - 1):
        texto_anterior = None
        inicio_col = None
        for col in range(1, max_col + 2):
            texto_actual = ws.cell(row=fila_encabezado + nivel, column=col).value if col <= max_col else None
            if texto_actual == texto_anterior and texto_actual is not None:
                pass
            else:
                if texto_anterior is not None and inicio_col is not None and (col - inicio_col) > 1:
                    rango = '{}{}:{}{}'.format(get_column_letter(inicio_col), fila_encabezado + nivel, get_column_letter(col - 1), fila_encabezado + nivel)
                    ws.merge_cells(rango)
                    ws.cell(row=fila_encabezado + nivel, column=inicio_col).alignment = Alignment(horizontal='center', vertical='center')
                texto_anterior = texto_actual
                inicio_col = col

    return fila_encabezado + max_segmentos - 1


def aplicar_formato_final(xlsx_path, json_path=None):
    reemplazos_encabezados, reemplazos_de_nombres = cargar_config_json(json_path)
    wb = load_workbook(xlsx_path)
    ws = wb.active

    base_name = os.path.splitext(os.path.basename(xlsx_path))[0]
    nombre_celda_a1 = reemplazos_de_nombres.get(base_name, 'Planilla')

    ws.cell(row=1, column=1, value=nombre_celda_a1)
    ws['A1'].font = FONT_TITULO
    ws['A1'].alignment = ALIGN_IZQ

    fila_encabezado = 6
    ultima_fila_encabezado = procesar_encabezados_divididos(ws, fila_encabezado, reemplazos_encabezados)

    fecha_hora_str = datetime.now().strftime('%d-%m-%y %H:%M')
    ws.cell(row=2, column=1, value='Fecha de generación: {}'.format(fecha_hora_str))
    ws['A2'].font = FONT_INFO
    ws['A2'].alignment = ALIGN_IZQ

    ws.cell(row=2, column=9).fill = RELLENO_GRIS
    ws.cell(row=3, column=9).fill = RELLENO_AZUL
    ws.cell(row=2, column=10, value='Columnas solo lectura').alignment = ALIGN_IZQ
    ws.cell(row=3, column=10, value='Columnas modificables').alignment = ALIGN_IZQ

    fila_inicio_datos = ultima_fila_encabezado + 1
    total_filas = 0
    for row in ws.iter_rows(min_row=fila_inicio_datos, max_row=ws.max_row):
        if any(cell.value not in (None, '') for cell in row):
            total_filas += 1

    ws.cell(row=3, column=1, value='Total Activos: {}'.format(total_filas))
    ws['A3'].font = FONT_INFO
    ws['A3'].alignment = ALIGN_IZQ

    wb.save(xlsx_path)


def convertir(csv_path, json_path=None):
    xlsx_path = os.path.splitext(csv_path)[0] + '.xlsx'
    crear_excel_base(csv_path, xlsx_path)
    aplicar_formato_final(xlsx_path, json_path)
    if os.path.exists(csv_path):
        os.remove(csv_path)
    return xlsx_path


if __name__ == '__main__':
    if len(sys.argv) < 2:
        sys.stderr.write('Debe indicar la ruta del CSV.\n')
        sys.exit(1)
    try:
        convertir(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else None)
    except Exception as exc:
        sys.stderr.write(str(exc))
        sys.exit(1)
